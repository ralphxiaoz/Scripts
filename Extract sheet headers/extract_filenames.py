import os
import pathlib
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

def extract_filenames(folder_paths, file_extensions=None, keywords=None):
    """
    Extract filenames from specified folders recursively with optional filtering
    
    Args:
        folder_paths (list): List of absolute paths to folders
        file_extensions (list): List of file extensions to filter by (e.g., ['.pdf', '.txt'])
        keywords (list): List of keywords that filenames must contain
        
    Returns:
        tuple: (list of tuples (full_path, filename, date_modified, source_folder), all_folder_paths)
    """
    all_file_data = []
    processed_folders = []
    
    for folder_path in folder_paths:
        try:
            # Convert to Path object for better path handling
            path = pathlib.Path(folder_path)
            
            # Check if path exists and is a directory
            if not path.exists():
                print(f"Warning: Path '{folder_path}' does not exist. Skipping...")
                continue
                
            if not path.is_dir():
                print(f"Warning: '{folder_path}' is not a directory. Skipping...")
                continue
            
            print(f"Processing folder: {folder_path}")
            processed_folders.append(folder_path)
            
            # Extract filenames recursively (including subdirectories)
            folder_file_count = 0
            
            # Use rglob to recursively find all files
            for item in path.rglob('*'):
                if item.is_file():
                    filename = item.name
                    full_path = str(item.parent)
                    
                    # Get file modification time
                    try:
                        mod_time = datetime.fromtimestamp(item.stat().st_mtime)
                        date_modified = mod_time.strftime('%Y-%m-%d %H:%M:%S')
                    except (OSError, ValueError):
                        date_modified = "Unknown"
                    
                    # Filter by file extensions if specified
                    if file_extensions:
                        file_ext = item.suffix.lower()
                        if file_ext not in [ext.lower() for ext in file_extensions]:
                            continue
                    
                    # Filter by keywords if specified
                    if keywords:
                        filename_lower = filename.lower()
                        if not any(keyword.lower() in filename_lower for keyword in keywords):
                            continue
                    
                    all_file_data.append((full_path, filename, date_modified, folder_path))
                    folder_file_count += 1
            
            print(f"  - Found {folder_file_count} files")
            
        except PermissionError:
            print(f"Warning: Permission denied to access '{folder_path}'. Skipping...")
            continue
        except Exception as e:
            print(f"Warning: Error processing '{folder_path}': {str(e)}. Skipping...")
            continue
    
    # Sort by source folder, then path, then filename
    all_file_data.sort(key=lambda x: (x[3].lower(), x[0].lower(), x[1].lower()))
    
    return all_file_data, processed_folders

def save_to_excel(file_data, folder_paths, output_path, file_extensions=None, keywords=None):
    """
    Save filenames to an Excel file with Path, Filename, Date Modified, and Source Folder columns
    
    Args:
        file_data (list): List of tuples (full_path, filename, date_modified, source_folder)
        folder_paths (list): List of original folder paths where search was performed
        output_path (str): Path to save the Excel file
        file_extensions (list): List of file extensions used for filtering
        keywords (list): List of keywords used for filtering
    """
    if not OPENPYXL_AVAILABLE:
        print("Error: openpyxl library is not installed. Please install it using:")
        print("pip install openpyxl")
        return
    
    try:
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted Filenames"
        
        # Set up headers
        ws['A1'] = 'Path'
        ws['B1'] = 'Filename'
        ws['C1'] = 'Date Modified'
        ws['D1'] = 'Source Folder'
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        for cell in ['A1', 'B1', 'C1', 'D1']:
            ws[cell].font = header_font
            ws[cell].fill = header_fill
        
        # Add data
        for i, (full_path, filename, date_modified, source_folder) in enumerate(file_data, start=2):
            ws[f'A{i}'] = full_path
            ws[f'B{i}'] = filename
            ws[f'C{i}'] = date_modified
            ws[f'D{i}'] = source_folder
        
        # Auto-adjust column widths
        max_path_len = max([len(str(row[0])) for row in file_data] + [50])
        ws.column_dimensions['A'].width = min(max_path_len + 5, 80)
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 40
        
        # Add metadata sheet
        metadata_ws = wb.create_sheet("Metadata")
        metadata_ws['A1'] = 'Extraction Details'
        metadata_ws['A1'].font = Font(bold=True, size=14)
        
        row = 3
        metadata_ws[f'A{row}'] = 'Extraction Date:'
        metadata_ws[f'B{row}'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        row += 1
        
        metadata_ws[f'A{row}'] = 'Source Folders:'
        metadata_ws[f'B{row}'] = f"{len(folder_paths)} folders processed"
        row += 1
        
        for i, folder_path in enumerate(folder_paths, 1):
            metadata_ws[f'A{row}'] = f'  Folder {i}:'
            metadata_ws[f'B{row}'] = folder_path
            row += 1
        
        metadata_ws[f'A{row}'] = 'Total Files Found:'
        metadata_ws[f'B{row}'] = len(file_data)
        row += 1
        
        if file_extensions:
            metadata_ws[f'A{row}'] = 'File Extensions Filter:'
            metadata_ws[f'B{row}'] = ', '.join(file_extensions)
            row += 1
            
        if keywords:
            metadata_ws[f'A{row}'] = 'Keywords Filter:'
            metadata_ws[f'B{row}'] = ', '.join(keywords)
            row += 1
        
        # Auto-adjust metadata column widths
        metadata_ws.column_dimensions['A'].width = 25
        metadata_ws.column_dimensions['B'].width = 50
        
        # Save the workbook
        wb.save(output_path)
        print(f"Excel file saved to: {output_path}")
        
    except Exception as e:
        print(f"Error saving Excel file: {str(e)}")

def save_to_file(file_data, output_path, file_extensions=None, keywords=None):
    """
    Save filenames to a text file
    
    Args:
        file_data (list): List of tuples (full_path, filename, date_modified, source_folder)
        output_path (str): Path to save the output file
        file_extensions (list): List of file extensions used for filtering
        keywords (list): List of keywords used for filtering
    """
    try:
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(f"Filenames extracted on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("=" * 50 + "\n")
            
            # Add filter information
            if file_extensions or keywords:
                f.write("Filters applied:\n")
                if file_extensions:
                    f.write(f"  - File extensions: {', '.join(file_extensions)}\n")
                if keywords:
                    f.write(f"  - Keywords: {', '.join(keywords)}\n")
                f.write("\n")
            
            f.write(f"Total files found: {len(file_data)}\n")
            f.write("-" * 30 + "\n\n")
            
            for full_path, filename, date_modified, source_folder in file_data:
                f.write(f"{full_path}\t{filename}\t{date_modified}\t{source_folder}\n")
                
        print(f"Filenames saved to: {output_path}")
        
    except Exception as e:
        print(f"Error saving file: {str(e)}")

def get_file_extensions():
    """
    Get file extensions from user input
    
    Returns:
        list or None: List of file extensions or None if no filtering
    """
    extensions_input = input("\nDo you want to filter by specific file extensions? (y/n): ").lower().strip()
    
    if extensions_input in ['y', 'yes']:
        while True:
            ext_input = input("Enter file extensions separated by commas (e.g., .pdf, .txt, .xlsx): ").strip()
            
            if not ext_input:
                print("Please enter at least one extension or press 'n' to skip filtering.")
                continue
            
            # Parse extensions
            extensions = [ext.strip() for ext in ext_input.split(',')]
            
            # Add dots if missing
            extensions = [ext if ext.startswith('.') else f'.{ext}' for ext in extensions]
            
            print(f"Will filter for files with extensions: {', '.join(extensions)}")
            return extensions
    
    return None

def get_keywords():
    """
    Get keywords from user input
    
    Returns:
        list or None: List of keywords or None if no filtering
    """
    keywords_input = input("\nDo you want to filter by keywords in filenames? (y/n): ").lower().strip()
    
    if keywords_input in ['y', 'yes']:
        while True:
            keyword_input = input("Enter keywords separated by commas (e.g., report, 2024, final): ").strip()
            
            if not keyword_input:
                print("Please enter at least one keyword or press 'n' to skip filtering.")
                continue
            
            # Parse keywords
            keywords = [keyword.strip() for keyword in keyword_input.split(',') if keyword.strip()]
            
            if not keywords:
                print("Please enter at least one valid keyword.")
                continue
            
            print(f"Will filter for files containing keywords: {', '.join(keywords)}")
            return keywords
    
    return None

def main():
    """
    Main function to run the filename extraction script
    """
    print("=== Filename Extractor ===")
    print("This script extracts filenames from a specified folder.\n")
    
    # Get folder paths from user
    while True:
        folder_input = input("Enter the absolute path(s) of the folder(s) (separate multiple paths with |): ").strip()
        
        if not folder_input:
            print("Please enter at least one valid path.")
            continue
            
        # Parse multiple folder paths using pipe delimiter
        folder_paths = [path.strip().strip('"\'') for path in folder_input.split('|')]
        folder_paths = [path for path in folder_paths if path]  # Remove empty strings
        
        if not folder_paths:
            print("Please enter at least one valid path.")
            continue
        
        # Display parsed paths for confirmation
        print(f"\nParsed {len(folder_paths)} folder path(s):")
        for i, path in enumerate(folder_paths, 1):
            print(f"  {i}. {path}")
        
        break
    
    # Get filtering options
    file_extensions = get_file_extensions()
    keywords = get_keywords()
    
    # Extract filenames
    print(f"\nExtracting filenames from {len(folder_paths)} folder(s)...")
    
    if file_extensions:
        print(f"Filtering by extensions: {', '.join(file_extensions)}")
    if keywords:
        print(f"Filtering by keywords: {', '.join(keywords)}")
    
    file_data, processed_folders = extract_filenames(folder_paths, file_extensions, keywords)
    
    if not file_data:
        print("No files found or error occurred.")
        return
    
    # Display results
    print(f"\nFound {len(file_data)} files across {len(processed_folders)} folders:")
    print("=" * 80)
    
    current_source = None
    for i, (full_path, filename, date_modified, source_folder) in enumerate(file_data, 1):
        # Group by source folder
        if source_folder != current_source:
            current_source = source_folder
            print(f"\nFrom folder: {source_folder}")
            print("-" * 60)
        
        # Show relative path for cleaner display
        try:
            rel_path = pathlib.Path(full_path).relative_to(pathlib.Path(source_folder))
            if str(rel_path) == ".":
                display_path = "."
            else:
                display_path = str(rel_path)
        except ValueError:
            # If relative path calculation fails, show full path
            display_path = full_path
            
        print(f"{i:3d}. {display_path}\\{filename} (Modified: {date_modified})")
    
    # Automatically save to Excel file
    print(f"\nTotal files found: {len(file_data)}")
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = f"filenames_{timestamp}.xlsx"
    
    if OPENPYXL_AVAILABLE:
        save_to_excel(file_data, processed_folders, output_path, file_extensions, keywords)
    else:
        print("Warning: openpyxl library is not installed. Cannot create Excel file.")
        print("Please install it using: pip install openpyxl")
        print("Saving as text file instead...")
        output_path = f"filenames_{timestamp}.txt"
        save_to_file(file_data, output_path, file_extensions, keywords)
    
    print("\nOperation completed!")

if __name__ == "__main__":
    main() 