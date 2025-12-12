import os
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime
import json
import xlrd
import csv
import fnmatch
import xml.etree.ElementTree as ET
import re
import yaml

def ensure_serializable(obj):
    """Convert objects to JSON-serializable format."""
    if isinstance(obj, datetime):
        return obj.strftime('%Y-%m-%d %H:%M:%S')
    elif isinstance(obj, dict):
        return {str(k): ensure_serializable(v) for k, v in obj.items()}
    elif isinstance(obj, (list, tuple)):
        return [ensure_serializable(item) for item in obj]
    elif isinstance(obj, (int, float, str, bool, type(None))):
        return obj
    else:
        return str(obj)

def unmerge_all_cells(sheet):
    """Unmerge all merged cells in a sheet."""
    try:
        # Get all merged cell ranges (need to convert to list to avoid modification during iteration)
        merged_ranges = list(sheet.merged_cells.ranges)
        for merged_range in merged_ranges:
            sheet.unmerge_cells(str(merged_range))
        return True
    except Exception as e:
        print(f"Warning: Could not unmerge cells: {e}")
        return False

def is_column_empty(sheet, col_idx, start_row=1, check_rows=100):
    """Check if a column is completely empty (all cells are None or empty string)."""
    for row in sheet.iter_rows(min_row=start_row, max_row=min(sheet.max_row, start_row + check_rows), 
                                min_col=col_idx, max_col=col_idx, values_only=True):
        if row[0] not in (None, ""):
            return False
    return True

def detect_header_row(sheet, scan_rows=20):
    """Detect the most likely header row in the first scan_rows of the sheet."""
    max_score = -1
    header_row_idx = 1
    
    # Get all rows to analyze
    all_rows = list(sheet.iter_rows(min_row=1, max_row=min(scan_rows + 5, sheet.max_row), values_only=True))
    
    for i, row in enumerate(all_rows[:scan_rows], 1):
        non_empty = [cell for cell in row if cell not in (None, "")]
        
        # Skip rows with too few columns (likely title rows)
        if len(non_empty) < 3:
            continue
            
        string_like = [cell for cell in non_empty if isinstance(cell, str)]
        unique = len(set(non_empty))
        
        # Base score
        score = len(non_empty) + len(string_like) + unique
        
        # Bonus: Check if followed by data rows with similar column structure
        if i < len(all_rows):
            next_rows_col_counts = []
            for j in range(i, min(i + 5, len(all_rows))):
                next_row_non_empty = [cell for cell in all_rows[j] if cell not in (None, "")]
                if len(next_row_non_empty) > 0:
                    next_rows_col_counts.append(len(next_row_non_empty))
            
            # If the next rows have similar column counts, boost score
            if next_rows_col_counts:
                avg_next_cols = sum(next_rows_col_counts) / len(next_rows_col_counts)
                if abs(avg_next_cols - len(non_empty)) <= 3:  # Within 3 columns
                    score += 10  # Bonus for data consistency
        
        if score > max_score:
            max_score = score
            header_row_idx = i
            
    return header_row_idx

def make_column_names_unique(columns):
    """Make column names unique by appending a number to duplicates."""
    seen = {}
    unique_columns = []
    for col in columns:
        if col in seen:
            seen[col] += 1
            unique_columns.append(f"{col}_{seen[col]}")
        else:
            seen[col] = 1
            unique_columns.append(col)
    return unique_columns

def get_column_formula_usage(sheet, header_row_idx, columns):
    # Returns a dict: {column_name: True/False} for formula usage
    formula_usage = {col: False for col in columns}
    max_cols = len(columns)
    for row in sheet.iter_rows(min_row=header_row_idx+1, max_row=sheet.max_row):
        for col_idx, col in enumerate(columns):
            if col_idx < len(row):
                cell = row[col_idx]
                if cell.data_type == 'f':
                    formula_usage[col] = True
    return formula_usage

def get_column_top_values_and_types(sheet, header_row_idx, columns, max_top=3):
    # Returns dict: {col: {'top_values': [...], 'empty_cells': (empty_count, total_count)}}
    data = list(sheet.iter_rows(min_row=header_row_idx+1, max_row=sheet.max_row, values_only=True))
    while data and all(cell in (None, "") for cell in data[-1]):
        data.pop()
    if not data:
        return {col: {'top_values': [], 'empty_cells': (0, 0)} for col in columns}
    max_cols = max(len(row) for row in data)
    if max_cols > len(columns):
        extra_columns = [f"Column_{i+1}" for i in range(len(columns), max_cols)]
        columns = columns + extra_columns
    unique_columns = make_column_names_unique(columns)
    df = pd.DataFrame(data, columns=unique_columns) if data else pd.DataFrame(columns=unique_columns)
    top_values = {}
    for orig_col, unique_col in zip(columns, unique_columns):
        col_data = df[unique_col]
        # Count empty/NA/NULL
        empty_mask = col_data.isna() | col_data.astype(str).str.strip().isin(["", "N/A", "NULL", "NA", "null", "n/a"])
        empty_count = int(empty_mask.sum())
        total_count = int(col_data.shape[0])
        col_data_nonan = col_data.dropna()
        if col_data_nonan.empty:
            top_values[orig_col] = {'top_values': [], 'empty_cells': (empty_count, total_count)}
            continue
        value_counts = col_data_nonan.value_counts().head(max_top)
        top_values[orig_col] = {
            'top_values': [str(val) for val in value_counts.index],
            'empty_cells': (empty_count, total_count)
        }
    return top_values

def extract_excel_from_xml(file_path, relative_path=None, scan_rows=20):
    """
    Extract data directly from .xlsx XML files, bypassing openpyxl entirely.
    This is the ultimate fallback for corrupted stylesheets.
    """
    import zipfile
    from xml.etree import ElementTree as ET
    
    try:
        file_name = os.path.basename(file_path)
        print(f"Extracting data directly from XML for {file_name}...")
        
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            # Read workbook.xml to get sheet names
            workbook_xml = zip_ref.read('xl/workbook.xml')
            workbook_tree = ET.fromstring(workbook_xml)
            
            # Namespace handling
            ns = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
                  'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}
            
            sheets = workbook_tree.findall('.//main:sheet', ns)
            if not sheets:  # Try without namespace
                sheets = workbook_tree.findall('.//sheet')
            
            metadata = {
                'file_name': file_name,
                'file_path': relative_path if relative_path else str(file_path),
                'total_sheets': len(sheets),
                'sheets': [],
                'note': 'Extracted via direct XML parsing (bypassing corrupted stylesheet)'
            }
            
            # Read shared strings if they exist
            shared_strings = []
            try:
                shared_strings_xml = zip_ref.read('xl/sharedStrings.xml')
                ss_tree = ET.fromstring(shared_strings_xml)
                for si in ss_tree.findall('.//main:si', ns):
                    t = si.find('.//main:t', ns)
                    if t is None:
                        t = si.find('.//t')
                    shared_strings.append(t.text if t is not None and t.text else '')
                if not shared_strings:  # Try without namespace
                    for si in ss_tree.findall('.//si'):
                        t = si.find('.//t')
                        shared_strings.append(t.text if t is not None and t.text else '')
            except KeyError:
                pass  # No shared strings
            
            for idx, sheet in enumerate(sheets, 1):
                sheet_name = sheet.get('name')
                sheet_id = sheet.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id') or sheet.get('r:id') or f'rId{idx}'
                
                try:
                    # Read the sheet XML
                    sheet_xml_path = f'xl/worksheets/sheet{idx}.xml'
                    sheet_xml = zip_ref.read(sheet_xml_path)
                    sheet_tree = ET.fromstring(sheet_xml)
                    
                    # Extract all rows
                    rows_data = []
                    for row in sheet_tree.findall('.//main:row', ns):
                        row_data = {}
                        for cell in row.findall('.//main:c', ns):
                            cell_ref = cell.get('r')
                            cell_type = cell.get('t')
                            v = cell.find('.//main:v', ns)
                            
                            if v is not None and v.text:
                                # Handle shared strings
                                if cell_type == 's':
                                    try:
                                        value = shared_strings[int(v.text)]
                                    except (IndexError, ValueError):
                                        value = v.text
                                else:
                                    value = v.text
                                
                                # Extract column from cell reference (e.g., 'A1' -> 'A')
                                col_letter = ''.join(filter(str.isalpha, cell_ref))
                                row_data[col_letter] = value
                        
                        if row_data:
                            rows_data.append(row_data)
                    
                    # Try without namespace if no data found
                    if not rows_data:
                        for row in sheet_tree.findall('.//row'):
                            row_data = {}
                            for cell in row.findall('.//c'):
                                cell_ref = cell.get('r')
                                cell_type = cell.get('t')
                                v = cell.find('.//v')
                                
                                if v is not None and v.text:
                                    if cell_type == 's':
                                        try:
                                            value = shared_strings[int(v.text)]
                                        except (IndexError, ValueError):
                                            value = v.text
                                    else:
                                        value = v.text
                                    
                                    col_letter = ''.join(filter(str.isalpha, cell_ref))
                                    row_data[col_letter] = value
                            
                            if row_data:
                                rows_data.append(row_data)
                    
                    if not rows_data:
                        metadata['sheets'].append({
                            'sheet_name': str(sheet_name),
                            'warning': 'sheet is empty'
                        })
                        continue
                    
                    # Convert rows_data to a list of lists for easier processing
                    # Get all column letters used
                    all_cols = set()
                    for row in rows_data:
                        all_cols.update(row.keys())
                    
                    # Sort column letters (A, B, C, ..., Z, AA, AB, ...)
                    def col_to_num(col):
                        num = 0
                        for char in col:
                            num = num * 26 + (ord(char) - ord('A') + 1)
                        return num
                    
                    sorted_cols = sorted(all_cols, key=col_to_num)
                    
                    # Convert to list of lists
                    grid_data = []
                    for row in rows_data:
                        grid_row = [row.get(col, '') for col in sorted_cols]
                        grid_data.append(grid_row)
                    
                    # Detect header row
                    header_row_idx = 0
                    max_score = -1
                    scan_limit = min(scan_rows, len(grid_data))
                    
                    for i in range(scan_limit):
                        row = grid_data[i]
                        non_empty = [str(cell) for cell in row if str(cell).strip() != '']
                        if not non_empty:
                            continue
                        score = len(non_empty) + len(set(non_empty))
                        text_count = sum(1 for cell in non_empty if not str(cell).replace('.', '', 1).replace('-', '', 1).isdigit())
                        score += text_count * 2
                        if score > max_score:
                            max_score = score
                            header_row_idx = i
                    
                    header_row = grid_data[header_row_idx] if header_row_idx < len(grid_data) else []
                    columns = [str(col).strip() if str(col).strip() else f"Column_{i+1}" 
                              for i, col in enumerate(header_row)]
                    
                    row_count = len(grid_data) - header_row_idx - 1
                    
                    metadata['sheets'].append({
                        'sheet_name': str(sheet_name),
                        'header_row': header_row_idx + 1,
                        'column_count': len(columns),
                        'columns': columns,
                        'row_count': row_count,
                        'note': 'Extracted via XML parsing'
                    })
                    
                except Exception as sheet_error:
                    print(f"Warning: Could not process sheet '{sheet_name}': {sheet_error}")
                    metadata['sheets'].append({
                        'sheet_name': str(sheet_name),
                        'error': str(sheet_error)
                    })
                    continue
            
            return metadata
            
    except Exception as e:
        return {
            'file_name': os.path.basename(file_path),
            'error': f"Direct XML extraction failed: {str(e)}"
        }

def extract_excel_metadata_pandas_fallback(file_path, relative_path=None, scan_rows=20):
    """Fallback method using pandas when openpyxl fails due to corrupted stylesheets."""
    try:
        print(f"Using pandas fallback method for {os.path.basename(file_path)} (corrupted stylesheet detected)")
        
        # Try different engines: first try xlrd for .xlsx, then calamine if available
        file_name = os.path.basename(file_path)
        xl_file = None
        engine_used = None
        
        # Try calamine engine first (fastest and most robust)
        try:
            xl_file = pd.ExcelFile(file_path, engine='calamine')
            engine_used = 'calamine'
        except:
            pass
        
        # If calamine doesn't work, we need another approach
        if xl_file is None:
            # Read using pandas with no engine specification (it will try its best)
            try:
                xl_file = pd.ExcelFile(file_path)
                engine_used = 'default'
            except:
                pass  # Will fall through to XML extraction
        
        # If all pandas engines failed, use direct XML extraction
        if xl_file is None:
            import zipfile
            from xml.etree import ElementTree as ET
            print(f"All pandas engines failed, attempting direct XML extraction from {file_name}...")
            return extract_excel_from_xml(file_path, relative_path, scan_rows)
        
        sheet_count = len(xl_file.sheet_names)
        print(f"Successfully opened with {engine_used} engine")
        
        metadata = {
            'file_name': file_name,
            'file_path': relative_path if relative_path else str(file_path),
            'total_sheets': sheet_count,
            'sheets': [],
            'note': 'Extracted using pandas fallback due to corrupted stylesheet - formula detection unavailable'
        }
        
        total_columns = 0
        
        for sheet_name in xl_file.sheet_names:
            try:
                # Read sheet with pandas - use whatever engine worked for xl_file
                df = pd.read_excel(xl_file, sheet_name=sheet_name, header=None)
                
                # Check if sheet is empty
                if df.empty or df.dropna(how='all').empty:
                    print(f"Warning: Sheet '{sheet_name}' is empty.")
                    metadata['sheets'].append({
                        'sheet_name': str(sheet_name),
                        'warning': 'sheet is empty'
                    })
                    continue
                
                # Remove completely empty rows at the end
                df = df.dropna(how='all')
                
                # Detect header row
                header_row_idx = 0
                max_score = -1
                scan_limit = min(scan_rows, len(df))
                
                for i in range(scan_limit):
                    row = df.iloc[i].fillna('')
                    non_empty = [str(cell) for cell in row if str(cell).strip() != '']
                    
                    if not non_empty:
                        continue
                    
                    # Score based on: non-numeric values, unique values, typical header patterns
                    score = 0
                    score += len(non_empty)  # More non-empty cells
                    score += len(set(non_empty))  # More unique values
                    
                    # Check for numeric vs text (headers are usually text)
                    text_count = sum(1 for cell in non_empty if not str(cell).replace('.', '', 1).replace('-', '', 1).isdigit())
                    score += text_count * 2
                    
                    if score > max_score:
                        max_score = score
                        header_row_idx = i
                
                # Extract header row
                header_row = df.iloc[header_row_idx].fillna('').tolist()
                
                # Extract columns
                columns = []
                for i, col in enumerate(header_row, 1):
                    col_str = str(col).strip()
                    if col_str and col_str != '':
                        columns.append(col_str)
                    else:
                        # Check if column has data
                        col_data = df.iloc[header_row_idx+1:, i-1].dropna()
                        if not col_data.empty:
                            columns.append(f"Column_{i}")
                
                if not columns:
                    print(f"Warning: Sheet '{sheet_name}' header missing.")
                    metadata['sheets'].append({
                        'sheet_name': str(sheet_name),
                        'warning': 'header missing, inherit from previous sheet'
                    })
                    continue
                
                # Get data rows (after header)
                data_df = df.iloc[header_row_idx+1:]
                row_count = len(data_df.dropna(how='all'))
                
                # Make columns unique
                unique_columns = make_column_names_unique(columns)
                
                # Build column details (simplified - no formula info)
                column_details = {}
                for i, col in enumerate(columns):
                    if i < len(data_df.columns):
                        col_data = data_df.iloc[:, i].dropna()
                        top_vals = col_data.value_counts().head(3).index.tolist()
                        
                        empty_count = data_df.iloc[:, i].isna().sum()
                        total_count = len(data_df)
                        
                        column_details[col] = {
                            'uses_formula': 'Unknown (pandas fallback)',
                            'top_values': [str(v) for v in top_vals],
                            'empty_cells (blank, N/A, NULL)': f"{empty_count}/{total_count}"
                        }
                
                total_columns += len(columns)
                
                sheet_info = {
                    'sheet_name': str(sheet_name),
                    'header_row': header_row_idx + 1,  # 1-indexed for Excel
                    'column_count': len(columns),
                    'columns': columns,
                    'unique_columns': unique_columns,
                    'row_count': row_count,
                    'skipped_rows': 0,
                    'has_pivot_table': False,
                    'column_details': column_details,
                    'original_header_count': len([col for col in header_row if str(col).strip() != '']),
                    'actual_data_columns': len(columns),
                    'duplicate_columns': {},
                    'sheet_formula_usage': 'N/A (pandas fallback)'
                }
                
                metadata['sheets'].append(sheet_info)
                
            except Exception as sheet_error:
                print(f"Warning: Could not process sheet '{sheet_name}': {sheet_error}")
                metadata['sheets'].append({
                    'sheet_name': str(sheet_name),
                    'error': str(sheet_error)
                })
                continue
        
        metadata['file_formula_usage'] = 'N/A (pandas fallback)'
        return metadata
        
    except Exception as e:
        return {
            'file_name': os.path.basename(file_path),
            'error': f"Pandas fallback also failed: {str(e)}"
        }

def extract_excel_metadata(file_path, relative_path=None, scan_rows=20):
    try:
        # First pass: get evaluated values (data_only=True)
        # Using keep_vba=False and rich_text=True for better compatibility with problematic files
        wb_values = load_workbook(file_path, data_only=True, keep_vba=False, rich_text=True)
        # Second pass: get formula info (data_only=False)
        wb_formulas = load_workbook(file_path, data_only=False, keep_vba=False, rich_text=True)
        file_name = os.path.basename(file_path)
        sheet_count = len(wb_values.sheetnames)
        metadata = {
            'file_name': file_name,
            'file_path': relative_path if relative_path else str(file_path),
            'total_sheets': sheet_count,
            'sheets': []
        }
        total_columns = 0
        total_formula_columns = 0
        for sheet_name in wb_values.sheetnames:
            sheet_values = wb_values[sheet_name]
            sheet_formulas = wb_formulas[sheet_name]
            
            # Unmerge all cells in both sheets
            unmerge_all_cells(sheet_values)
            unmerge_all_cells(sheet_formulas)
            
            # Check if the entire sheet is empty before looking for headers
            all_data_rows = list(sheet_values.iter_rows(min_row=1, max_row=sheet_values.max_row, values_only=True))
            while all_data_rows and all(cell in (None, "") for cell in all_data_rows[-1]):
                all_data_rows.pop()
            
            if not all_data_rows or all(all(cell in (None, "") for cell in row) for row in all_data_rows):
                print(f"Warning: Sheet '{sheet_name}' is empty.")
                metadata['sheets'].append({
                    'sheet_name': str(sheet_name),
                    'warning': 'sheet is empty'
                })
                continue
            
            header_row_idx = detect_header_row(sheet_values, scan_rows)
            header_row = [cell.value for cell in next(sheet_values.iter_rows(min_row=header_row_idx, max_row=header_row_idx))]
            
            # First, extract non-empty headers for validation
            non_empty_headers = [col for col in header_row if col is not None and str(col).strip() != ""]
            
            # Validate that we have a likely header (using only non-empty values)
            if not non_empty_headers or not is_likely_header(non_empty_headers):
                print(f"Warning: Sheet '{sheet_name}' header missing, inherit from previous sheet.")
                metadata['sheets'].append({
                    'sheet_name': str(sheet_name),
                    'warning': 'header missing, inherit from previous sheet'
                })
                continue
            
            # Extract columns, skipping completely empty columns (from merged cells)
            columns = []
            for i, col in enumerate(header_row, 1):
                # If header cell is empty, check if column has data
                if col is None or str(col).strip() == "":
                    # Check if this column has any data below the header
                    if not is_column_empty(sheet_values, i, start_row=header_row_idx+1, check_rows=100):
                        # Column has data but no header, generate a name
                        columns.append(f"Column_{i}")
                    # Otherwise skip this empty column
                else:
                    # Header has a value, keep it
                    columns.append(str(col).strip())
            
            # Remove trailing auto-generated column names if they're truly empty
            while columns and columns[-1].startswith("Column_"):
                columns.pop()
                
            data_rows = list(sheet_values.iter_rows(min_row=header_row_idx+1, max_row=sheet_values.max_row, values_only=True))
            while data_rows and all(cell in (None, "") for cell in data_rows[-1]):
                data_rows.pop()
            
            max_cols = max(len(row) for row in data_rows) if data_rows else 0
            if max_cols > len(columns):
                extra_columns = [f"Column_{i+1}" for i in range(len(columns), max_cols)]
                columns = columns + extra_columns
            unique_columns = make_column_names_unique(columns)
            row_count = len(data_rows)
            skipped_rows = 0
            found_data = False
            for row in data_rows:
                if all(cell in (None, "") for cell in row):
                    if found_data:
                        skipped_rows += 1
                else:
                    found_data = True
            has_pivot = False
            for pivot in getattr(sheet_values, '_pivots', []):
                has_pivot = True
                break
            # Get top values/types (evaluated values)
            top_values_types = get_column_top_values_and_types(sheet_values, header_row_idx, columns, max_top=3)
            # Get formula usage (from formula sheet)
            formula_usage = get_column_formula_usage(sheet_formulas, header_row_idx, columns)
            # Count formula columns
            formula_columns = sum(1 for col in columns if formula_usage.get(col, False))
            total_columns += len(columns)
            total_formula_columns += formula_columns
            duplicate_columns = {}
            seen = {}
            for i, col in enumerate(columns):
                if col in seen:
                    if col not in duplicate_columns:
                        duplicate_columns[col] = [seen[col]]
                    duplicate_columns[col].append(i)
                else:
                    seen[col] = i
            # Build column_details
            column_details = {}
            for col in columns:
                col_info = {
                    'uses_formula': formula_usage.get(col, False)
                }
                if 'top_values' in top_values_types[col] and top_values_types[col]['top_values']:
                    col_info['top_values'] = top_values_types[col]['top_values']
                if 'empty_cells' in top_values_types[col]:
                    empty_count, total_count = top_values_types[col]['empty_cells']
                    col_info['empty_cells (blank, N/A, NULL)'] = f"{empty_count}/{total_count}"
                column_details[col] = col_info
            sheet_info = {
                'sheet_name': str(sheet_name),
                'header_row': header_row_idx,
                'column_count': len(columns),
                'columns': columns,
                'unique_columns': unique_columns,
                'row_count': row_count,
                'skipped_rows': skipped_rows,
                'has_pivot_table': has_pivot,
                'column_details': column_details,
                'original_header_count': len([col for col in header_row if col not in (None, "")]),
                'actual_data_columns': max_cols,
                'duplicate_columns': duplicate_columns,
                'sheet_formula_usage': f"{formula_columns}/{len(columns)} columns"
            }
            if getattr(sheet_values, 'protection', None) and sheet_values.protection.sheet:
                print(f"Warning: Sheet '{sheet_name}' is protected. Skipping or reporting limited metadata.")
                metadata['sheets'].append({
                    'sheet_name': str(sheet_name),
                    'warning': 'Sheet is protected and was skipped from metadata extraction.'
                })
                continue
            metadata['sheets'].append(sheet_info)
        metadata['file_formula_usage'] = f"{total_formula_columns}/{total_columns} columns"
        return metadata
    except Exception as e:
        error_str = str(e)
        # Check if this is a stylesheet/XML corruption error
        if 'stylesheet' in error_str.lower() or 'invalid xml' in error_str.lower():
            print(f"Stylesheet corruption detected in {os.path.basename(file_path)}, trying pandas fallback...")
            return extract_excel_metadata_pandas_fallback(file_path, relative_path, scan_rows)
        else:
            return {
                'file_name': os.path.basename(file_path),
                'error': str(e)
            }

def format_metadata(metadata):
    """Format metadata as a JSON-compatible dictionary."""
    if 'error' in metadata:
        return {
            'file_name': metadata['file_name'],
            'error': metadata['error']
        }
    
    formatted = {
        'file_name': metadata['file_name'],
        'total_sheets': metadata.get('total_sheets', 1),
        'formula_usage': metadata.get('file_formula_usage', 'N/A'),
        'sheets': []
    }
    
    for sheet in metadata['sheets']:
        if 'warning' in sheet:
            # Just include the warning and sheet name
            formatted['sheets'].append({
                'sheet_name': sheet['sheet_name'],
                'warning': sheet['warning']
            })
            continue
        
        # Check if this is a simple sheet (like unsupported files) with minimal info
        if 'header_row' not in sheet:
            # This is a minimal sheet (e.g., unsupported file type)
            formatted['sheets'].append({
                'sheet_name': sheet.get('sheet_name', 'N/A'),
                'columns': sheet.get('columns', [])
            })
            continue
        
        # Full sheet with all metadata
        sheet_info = {
            'sheet_name': sheet['sheet_name'],
            'header_row': sheet['header_row'],
            'column_count': sheet['column_count'],
            'columns': sheet['columns'],
            'row_count': sheet['row_count'],
            'skipped_rows': sheet['skipped_rows'],
            'has_pivot_table': sheet['has_pivot_table'],
            'formula_usage': sheet.get('sheet_formula_usage', None),
            'column_details': {}
        }
        
        for col in sheet['columns']:
            top = sheet['column_details'].get(col, {})
            col_detail = {
                'uses_formula': top.get('uses_formula', False)
            }
            if top.get('top_values'):
                col_detail['top_values'] = top['top_values']
            if top.get('empty_cells (blank, N/A, NULL)'):
                col_detail['empty_cells (blank, N/A, NULL)'] = top['empty_cells (blank, N/A, NULL)']
            sheet_info['column_details'][col] = col_detail
        
        formatted['sheets'].append(sheet_info)
    
    return formatted

def get_folder_selection():
    """Get user selection for which folder to scan."""
    while True:
        try:
            target_path = input("\nEnter the absolute path of the folder to scan: ").strip()
            # Remove quotes if user included them
            target_path = target_path.strip('"\'')
            
            # Convert to Path object
            target_folder = Path(target_path)
            
            if not target_folder.exists():
                print(f"Error: The path '{target_path}' does not exist.")
                continue
                
            if not target_folder.is_dir():
                print(f"Error: '{target_path}' is not a directory.")
                continue
            
            # Check if there are subfolders in the target directory (only immediate subfolders for display)
            subfolders = [f for f in target_folder.iterdir() if f.is_dir()]
            
            if not subfolders:
                print(f"\nNo subfolders found in '{target_folder.name}'. Will scan the folder recursively.")
                return [target_folder]
            
            print(f"\nFound the following immediate subfolders in '{target_folder.name}':")
            print("-" * 50)
            for i, folder in enumerate(subfolders, 1):
                print(f"{i}. {folder.name}")
            print("0. Scan all subfolders recursively (includes root folder)")
            print(f"T. Scan the target folder '{target_folder.name}' recursively")
            
            while True:
                try:
                    choice = input("\nEnter your choice (0 for all subfolders, T for target folder, or number for specific subfolder): ").strip()
                    
                    if choice.upper() == 'T':
                        return [target_folder]
                    elif choice == '0':
                        # Include the target folder AND all subfolders
                        return [target_folder] + subfolders
                    else:
                        choice_num = int(choice)
                        if 1 <= choice_num <= len(subfolders):
                            return [subfolders[choice_num - 1]]
                        else:
                            print(f"Please enter a number between 1 and {len(subfolders)}, 0 for all, or T for target folder")
                except ValueError:
                    print("Please enter a valid number, 0 for all, or T for target folder")
                    
        except Exception as e:
            print(f"Error processing the path: {e}")
            print("Please enter a valid absolute path.")

def get_user_selection(excel_files):
    print("\nFound the following files:")
    print("-" * 50)
    for i, file in enumerate(excel_files, 1):
        # Show relative path to make it clearer which subfolder the file is in
        print(f"{i}. {file}")
    print("0. Analyze all files")
    while True:
        try:
            choice = input("\nEnter the number of the file to analyze (0 for all): ")
            choice = int(choice)
            if 0 <= choice <= len(excel_files):
                return choice
            print(f"Please enter a number between 0 and {len(excel_files)}")
        except ValueError:
            print("Please enter a valid number")

def get_user_mode():
    while True:
        mode = input("Choose extraction mode: [1] Simple (file/sheet/column names only), [2] Full (detailed metadata): ").strip()
        if mode in ('1', 'simple', 'Simple'):
            return 'simple'
        elif mode in ('2', 'full', 'Full'):
            return 'full'
        else:
            print("Please enter 1 for Simple or 2 for Full mode.")

def get_excel_output_choice():
    """Ask user if they want an Excel file in addition to JSON."""
    while True:
        choice = input("Do you also want an Excel (.xlsx) file with the results? (y/n): ").strip().lower()
        if choice in ['y', 'yes']:
            return True
        elif choice in ['n', 'no']:
            return False
        else:
            print("Please enter 'y' or 'n'.")

def convert_json_to_excel(json_file_path):
    """Convert the JSON output to Excel format with file_path, file_name, sheet, and field_name columns."""
    try:
        # Read the JSON file
        with open(json_file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        rows = []
        for file_data in data['files']:
            file_name = file_data['file_name']
            file_path = file_data.get('file_path', file_name)  # Default to file_name if no file_path
            
            # Handle error cases
            if 'error' in file_data:
                rows.append({
                    'file_path': file_path,
                    'file_name': file_name,
                    'sheet': 'ERROR',
                    'field_name': f"ERROR: {file_data['error']}"
                })
                continue
            
            # Handle files with sheets
            if 'sheets' in file_data:
                for sheet in file_data['sheets']:
                    sheet_name = sheet.get('sheet_name', 'Unknown')
                    
                    # Handle warning cases (empty sheets, protected sheets, etc.)
                    if 'warning' in sheet:
                        warning_msg = sheet['warning']
                        # Check if it's a header missing warning
                        if 'header missing' in warning_msg:
                            rows.append({
                                'file_path': file_path,
                                'file_name': file_name,
                                'sheet': sheet_name,
                                'field_name': f"WARNING: No header detected - {warning_msg}"
                            })
                        else:
                            rows.append({
                                'file_path': file_path,
                                'file_name': file_name,
                                'sheet': sheet_name,
                                'field_name': f"WARNING: {warning_msg}"
                            })
                        continue
                    
                    # Handle normal sheets with columns
                    if 'columns' in sheet:
                        for column in sheet['columns']:
                            rows.append({
                                'file_path': file_path,
                                'file_name': file_name,
                                'sheet': sheet_name,
                                'field_name': column
                            })
                    
                    # If no columns but sheet exists, add a note
                    if 'columns' not in sheet and 'warning' not in sheet:
                        rows.append({
                            'file_path': file_path,
                            'file_name': file_name,
                            'sheet': sheet_name,
                            'field_name': "WARNING: No columns found"
                        })
        
        # Create Excel file with same name but .xlsx extension
        excel_file_path = json_file_path.replace('.json', '.xlsx')
        df = pd.DataFrame(rows)
        df.to_excel(excel_file_path, index=False)
        return excel_file_path
        
    except Exception as e:
        print(f"Error converting JSON to Excel: {e}")
        return None

def extract_excel_simple_pandas_fallback(file_path, relative_path=None):
    """Fallback simple extraction using pandas when openpyxl fails."""
    try:
        print(f"Using pandas fallback for simple extraction of {os.path.basename(file_path)}")
        file_name = os.path.basename(file_path)
        xl_file = None
        
        # Try calamine engine first
        try:
            xl_file = pd.ExcelFile(file_path, engine='calamine')
        except:
            pass
        
        # Try default engine
        if xl_file is None:
            try:
                xl_file = pd.ExcelFile(file_path)
            except:
                pass
        
        # If all pandas engines failed, use XML extraction
        if xl_file is None:
            print(f"All pandas engines failed, using direct XML extraction for {file_name}...")
            return extract_excel_from_xml(file_path, relative_path, 20)
        
        simple_metadata = {
            'file_name': file_name,
            'file_path': relative_path if relative_path else str(file_path),
            'sheets': [],
            'note': 'Extracted using pandas fallback due to corrupted stylesheet'
        }
        
        for sheet_name in xl_file.sheet_names:
            try:
                # Don't specify engine - use whatever worked for xl_file
                df = pd.read_excel(xl_file, sheet_name=sheet_name, header=None)
                
                if df.empty or df.dropna(how='all').empty:
                    simple_metadata['sheets'].append({
                        'sheet_name': str(sheet_name),
                        'warning': 'sheet is empty'
                    })
                    continue
                
                df = df.dropna(how='all')
                
                # Detect header row
                header_row_idx = 0
                max_score = -1
                for i in range(min(20, len(df))):
                    row = df.iloc[i].fillna('')
                    non_empty = [str(cell) for cell in row if str(cell).strip() != '']
                    if not non_empty:
                        continue
                    score = len(non_empty) + len(set(non_empty))
                    text_count = sum(1 for cell in non_empty if not str(cell).replace('.', '', 1).replace('-', '', 1).isdigit())
                    score += text_count * 2
                    if score > max_score:
                        max_score = score
                        header_row_idx = i
                
                header_row = df.iloc[header_row_idx].fillna('').tolist()
                columns = []
                for i, col in enumerate(header_row, 1):
                    col_str = str(col).strip()
                    if col_str and col_str != '':
                        columns.append(col_str)
                    else:
                        col_data = df.iloc[header_row_idx+1:, i-1].dropna()
                        if not col_data.empty:
                            columns.append(f"Column_{i}")
                
                if not columns:
                    simple_metadata['sheets'].append({
                        'sheet_name': str(sheet_name),
                        'warning': 'header missing, inherit from previous headers'
                    })
                    continue
                
                simple_metadata['sheets'].append({
                    'sheet_name': str(sheet_name),
                    'columns': columns
                })
                
            except Exception as sheet_error:
                print(f"Warning: Could not process sheet '{sheet_name}': {sheet_error}")
                simple_metadata['sheets'].append({
                    'sheet_name': str(sheet_name),
                    'error': str(sheet_error)
                })
                continue
        
        return simple_metadata
        
    except Exception as e:
        return {
            'file_name': os.path.basename(file_path),
            'error': f"Pandas fallback also failed: {str(e)}"
        }

def extract_excel_simple(file_path, relative_path=None):
    try:
        wb = load_workbook(file_path, data_only=True, keep_vba=False, rich_text=True)
        file_name = os.path.basename(file_path)
        simple_metadata = {
            'file_name': file_name,
            'file_path': relative_path if relative_path else str(file_path),
            'sheets': []
        }
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            
            # Unmerge all cells first
            unmerge_all_cells(sheet)
            
            if getattr(sheet, 'protection', None) and sheet.protection.sheet:
                print(f"Warning: Sheet '{sheet_name}' is protected. Skipping or reporting limited metadata.")
                simple_metadata['sheets'].append({
                    'sheet_name': str(sheet_name),
                    'warning': 'Sheet is protected and was skipped from metadata extraction.'
                })
                continue
            
            # Check if the entire sheet is empty before looking for headers
            all_data_rows = list(sheet.iter_rows(min_row=1, max_row=sheet.max_row, values_only=True))
            while all_data_rows and all(cell in (None, "") for cell in all_data_rows[-1]):
                all_data_rows.pop()
            
            if not all_data_rows or all(all(cell in (None, "") for cell in row) for row in all_data_rows):
                print(f"Warning: Sheet '{sheet_name}' is empty.")
                simple_metadata['sheets'].append({
                    'sheet_name': str(sheet_name),
                    'warning': 'sheet is empty'
                })
                continue
            
            header_row_idx = detect_header_row(sheet, scan_rows=20)
            header_row = [cell.value for cell in next(sheet.iter_rows(min_row=header_row_idx, max_row=header_row_idx))]
            
            # First, extract non-empty headers for validation
            non_empty_headers = [col for col in header_row if col is not None and str(col).strip() != ""]
            
            # Validate that we have a likely header (using only non-empty values)
            if not non_empty_headers or not is_likely_header(non_empty_headers):
                print(f"Warning: Sheet '{sheet_name}' header missing, inherit from previous headers.")
                simple_metadata['sheets'].append({
                    'sheet_name': str(sheet_name),
                    'warning': 'header missing, inherit from previous headers'
                })
                continue
            
            # Extract columns, skipping completely empty columns (from merged cells)
            columns = []
            for i, col in enumerate(header_row, 1):
                # If header cell is empty, check if column has data
                if col is None or str(col).strip() == "":
                    # Check if this column has any data below the header
                    if not is_column_empty(sheet, i, start_row=header_row_idx+1, check_rows=100):
                        # Column has data but no header, generate a name
                        columns.append(f"Column_{i}")
                    # Otherwise skip this empty column
                else:
                    # Header has a value, keep it
                    columns.append(str(col).strip())
            
            # Remove trailing auto-generated column names if they're truly empty
            while columns and columns[-1].startswith("Column_"):
                columns.pop()
            

            
            simple_metadata['sheets'].append({
                'sheet_name': str(sheet_name),
                'columns': columns
            })
        return simple_metadata
    except Exception as e:
        error_str = str(e)
        # Check if this is a stylesheet/XML corruption error
        if 'stylesheet' in error_str.lower() or 'invalid xml' in error_str.lower():
            print(f"Stylesheet corruption detected in {os.path.basename(file_path)}, trying pandas fallback...")
            return extract_excel_simple_pandas_fallback(file_path, relative_path)
        else:
            return {
                'file_name': os.path.basename(file_path),
                'error': str(e)
            }

def is_xls_column_empty(sheet, col_idx, start_row=0, check_rows=100):
    """Check if a column in an XLS sheet is completely empty."""
    end_row = min(sheet.nrows, start_row + check_rows)
    for row_idx in range(start_row, end_row):
        try:
            cell_value = sheet.cell_value(row_idx, col_idx)
            if cell_value not in (None, "", xlrd.empty_cell.value):
                return False
        except IndexError:
            # Column doesn't exist at this row
            pass
    return True

def extract_xls_metadata(file_path, relative_path=None, scan_rows=20):
    """Extract metadata from .xls files using xlrd."""
    try:
        wb = xlrd.open_workbook(file_path, formatting_info=False)
        file_name = os.path.basename(file_path)
        sheet_count = len(wb.sheet_names())
        metadata = {
            'file_name': file_name,
            'file_path': relative_path if relative_path else str(file_path),
            'total_sheets': sheet_count,
            'sheets': []
        }
        total_columns = 0
        
        for sheet_name in wb.sheet_names():
            sheet = wb.sheet_by_name(sheet_name)
            
            # Check if the entire sheet is empty before looking for headers
            all_data_rows = []
            for i in range(sheet.nrows):
                row = sheet.row_values(i)
                all_data_rows.append(row)
            
            # Remove empty rows from end
            while all_data_rows and all(cell in (None, "", xlrd.empty_cell.value) for cell in all_data_rows[-1]):
                all_data_rows.pop()
            
            if not all_data_rows or all(all(cell in (None, "", xlrd.empty_cell.value) for cell in row) for row in all_data_rows):
                print(f"Warning: Sheet '{sheet_name}' is empty.")
                metadata['sheets'].append({
                    'sheet_name': str(sheet_name),
                    'warning': 'sheet is empty'
                })
                continue
            
            # Detect header row with improved logic
            header_row_idx = 0
            max_score = -1
            scan_limit = min(scan_rows, sheet.nrows)
            
            for i in range(scan_limit):
                row = sheet.row_values(i)
                non_empty = [cell for cell in row if cell not in (None, "", xlrd.empty_cell.value)]
                
                # Skip rows with too few columns (likely title rows)
                if len(non_empty) < 3:
                    continue
                
                string_like = [cell for cell in non_empty if isinstance(cell, str)]
                unique = len(set(non_empty))
                
                # Base score
                score = len(non_empty) + len(string_like) + unique
                
                # Bonus: Check if followed by data rows with similar column structure
                if i < sheet.nrows - 1:
                    next_rows_col_counts = []
                    for j in range(i + 1, min(i + 6, sheet.nrows)):
                        next_row = sheet.row_values(j)
                        next_row_non_empty = [cell for cell in next_row if cell not in (None, "", xlrd.empty_cell.value)]
                        if len(next_row_non_empty) > 0:
                            next_rows_col_counts.append(len(next_row_non_empty))
                    
                    # If the next rows have similar column counts, boost score
                    if next_rows_col_counts:
                        avg_next_cols = sum(next_rows_col_counts) / len(next_rows_col_counts)
                        if abs(avg_next_cols - len(non_empty)) <= 3:  # Within 3 columns
                            score += 10  # Bonus for data consistency
                
                if score > max_score:
                    max_score = score
                    header_row_idx = i
            
            # Get header row and extract columns, skipping empty ones
            header_row = sheet.row_values(header_row_idx)
            
            # First, extract the non-empty header values for validation
            non_empty_headers = [col for col in header_row if col not in (None, "", xlrd.empty_cell.value)]
            
            # Validate that we have a likely header (using only non-empty values)
            if not non_empty_headers or not is_likely_header(non_empty_headers):
                print(f"Warning: Sheet '{sheet_name}' header missing, inherit from previous sheet.")
                metadata['sheets'].append({
                    'sheet_name': str(sheet_name),
                    'warning': 'header missing, inherit from previous sheet'
                })
                continue
            
            # Now extract all columns, including empty ones that have data
            columns = []
            for col_idx, col_value in enumerate(header_row):
                # If header cell is empty, check if column has data
                if col_value in (None, "", xlrd.empty_cell.value):
                    # Check if this column has any data below the header
                    if not is_xls_column_empty(sheet, col_idx, start_row=header_row_idx+1, check_rows=100):
                        # Column has data but no header, generate a name
                        columns.append(f"Column_{col_idx+1}")
                    # Otherwise skip this empty column
                else:
                    # Header has a value, keep it
                    columns.append(str(col_value))
            
            # Remove trailing auto-generated column names if they're truly empty
            while columns and columns[-1].startswith("Column_"):
                columns.pop()
            
            # Get data rows
            data_rows = []
            for i in range(header_row_idx + 1, sheet.nrows):
                row = sheet.row_values(i)
                data_rows.append(row)
            
            # Remove empty rows from end
            while data_rows and all(cell in (None, "", xlrd.empty_cell.value) for cell in data_rows[-1]):
                data_rows.pop()
            
            max_cols = max(len(row) for row in data_rows) if data_rows else 0
            if max_cols > len(columns):
                extra_columns = [f"Column_{i+1}" for i in range(len(columns), max_cols)]
                columns = columns + extra_columns
            
            unique_columns = make_column_names_unique(columns)
            row_count = len(data_rows)
            
            # Count skipped rows
            skipped_rows = 0
            found_data = False
            for row in data_rows:
                if all(cell in (None, "", xlrd.empty_cell.value) for cell in row):
                    if found_data:
                        skipped_rows += 1
                else:
                    found_data = True
            
            # Get top values and types (simplified for xlrd)
            top_values_types = {}
            for col_idx, col in enumerate(columns):
                col_data = []
                for row in data_rows:
                    if col_idx < len(row):
                        cell_value = row[col_idx]
                        if cell_value not in (None, "", xlrd.empty_cell.value):
                            col_data.append(str(cell_value))
                
                empty_count = sum(1 for row in data_rows 
                                if col_idx >= len(row) or row[col_idx] in (None, "", xlrd.empty_cell.value))
                total_count = len(data_rows)
                
                # Get top values
                value_counts = {}
                for val in col_data:
                    value_counts[val] = value_counts.get(val, 0) + 1
                
                top_values = sorted(value_counts.items(), key=lambda x: x[1], reverse=True)[:3]
                top_values_types[col] = {
                    'top_values': [val for val, count in top_values],
                    'empty_cells': (empty_count, total_count)
                }
            
            # Build column details
            column_details = {}
            for col in columns:
                col_info = {
                    'uses_formula': False  # xlrd doesn't easily provide formula info
                }
                if 'top_values' in top_values_types[col] and top_values_types[col]['top_values']:
                    col_info['top_values'] = top_values_types[col]['top_values']
                if 'empty_cells' in top_values_types[col]:
                    empty_count, total_count = top_values_types[col]['empty_cells']
                    col_info['empty_cells (blank, N/A, NULL)'] = f"{empty_count}/{total_count}"
                column_details[col] = col_info
            
            # Check for duplicate columns
            duplicate_columns = {}
            seen = {}
            for i, col in enumerate(columns):
                if col in seen:
                    if col not in duplicate_columns:
                        duplicate_columns[col] = [seen[col]]
                    duplicate_columns[col].append(i)
                else:
                    seen[col] = i
            
            sheet_info = {
                'sheet_name': str(sheet_name),
                'header_row': header_row_idx + 1,  # Convert to 1-based indexing
                'column_count': len(columns),
                'columns': columns,
                'unique_columns': unique_columns,
                'row_count': row_count,
                'skipped_rows': skipped_rows,
                'has_pivot_table': False,  # xlrd doesn't easily detect pivot tables
                'column_details': column_details,
                'original_header_count': len([col for col in header_row if col not in (None, "", xlrd.empty_cell.value)]),
                'actual_data_columns': max_cols,
                'duplicate_columns': duplicate_columns,
                'sheet_formula_usage': f"0/{len(columns)} columns"  # xlrd doesn't provide formula info
            }
            metadata['sheets'].append(sheet_info)
            total_columns += len(columns)
        
        metadata['file_formula_usage'] = f"0/{total_columns} columns"
        return metadata
        
    except Exception as e:
        return {
            'file_name': os.path.basename(file_path),
            'error': str(e)
        }

def extract_xls_simple(file_path, relative_path=None):
    """Extract simple metadata from .xls files using xlrd."""
    try:
        wb = xlrd.open_workbook(file_path)
        file_name = os.path.basename(file_path)
        simple_metadata = {
            'file_name': file_name,
            'file_path': relative_path if relative_path else str(file_path),
            'sheets': []
        }
        for sheet_name in wb.sheet_names():
            sheet = wb.sheet_by_name(sheet_name)
            
            # Check if the entire sheet is empty before looking for headers
            all_data_rows = []
            for i in range(sheet.nrows):
                row = sheet.row_values(i)
                all_data_rows.append(row)
            
            # Remove empty rows from end
            while all_data_rows and all(cell in (None, "", xlrd.empty_cell.value) for cell in all_data_rows[-1]):
                all_data_rows.pop()
            
            if not all_data_rows or all(all(cell in (None, "", xlrd.empty_cell.value) for cell in row) for row in all_data_rows):
                print(f"Warning: Sheet '{sheet_name}' is empty.")
                simple_metadata['sheets'].append({
                    'sheet_name': str(sheet_name),
                    'warning': 'sheet is empty'
                })
                continue
            
            # Detect header row with improved logic
            header_row_idx = 0
            max_score = -1
            scan_limit = min(20, sheet.nrows)
            
            for i in range(scan_limit):
                row = sheet.row_values(i)
                non_empty = [cell for cell in row if cell not in (None, "", xlrd.empty_cell.value)]
                
                # Skip rows with too few columns (likely title rows)
                if len(non_empty) < 3:
                    continue
                
                string_like = [cell for cell in non_empty if isinstance(cell, str)]
                unique = len(set(non_empty))
                
                # Base score
                score = len(non_empty) + len(string_like) + unique
                
                # Bonus: Check if followed by data rows with similar column structure
                if i < sheet.nrows - 1:
                    next_rows_col_counts = []
                    for j in range(i + 1, min(i + 6, sheet.nrows)):
                        next_row = sheet.row_values(j)
                        next_row_non_empty = [cell for cell in next_row if cell not in (None, "", xlrd.empty_cell.value)]
                        if len(next_row_non_empty) > 0:
                            next_rows_col_counts.append(len(next_row_non_empty))
                    
                    # If the next rows have similar column counts, boost score
                    if next_rows_col_counts:
                        avg_next_cols = sum(next_rows_col_counts) / len(next_rows_col_counts)
                        if abs(avg_next_cols - len(non_empty)) <= 3:  # Within 3 columns
                            score += 10  # Bonus for data consistency
                
                if score > max_score:
                    max_score = score
                    header_row_idx = i
            # Get header row
            header_row = sheet.row_values(header_row_idx)
            
            # First, extract the non-empty header values for validation
            non_empty_headers = [col for col in header_row if col not in (None, "", xlrd.empty_cell.value)]
            
            # Validate that we have a likely header (using only non-empty values)
            if not non_empty_headers or not is_likely_header(non_empty_headers):
                print(f"Warning: Sheet '{sheet_name}' header missing, inherit from previous headers.")
                simple_metadata['sheets'].append({
                    'sheet_name': str(sheet_name),
                    'warning': 'header missing, inherit from previous headers'
                })
                continue
            
            # Now extract all columns, skipping truly empty ones
            columns = []
            for col_idx, col_value in enumerate(header_row):
                # If header cell is empty, check if column has data
                if col_value in (None, "", xlrd.empty_cell.value):
                    # Check if this column has any data below the header
                    if not is_xls_column_empty(sheet, col_idx, start_row=header_row_idx+1, check_rows=100):
                        # Column has data but no header, generate a name
                        columns.append(f"Column_{col_idx+1}")
                    # Otherwise skip this empty column
                else:
                    # Header has a value, keep it
                    columns.append(str(col_value))
            
            # Remove trailing auto-generated column names if they're truly empty
            while columns and columns[-1].startswith("Column_"):
                columns.pop()
            
            simple_metadata['sheets'].append({
                'sheet_name': str(sheet_name),
                'columns': columns
            })
        return simple_metadata
    except Exception as e:
        return {
            'file_name': os.path.basename(file_path),
            'error': str(e)
        }

def read_txt_pipe_delimited(file_path, scan_rows=10):
    """Read pipe-delimited text file and return rows."""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            lines = [line.rstrip('\n\r') for line in f]
            return lines
    except UnicodeDecodeError:
        with open(file_path, 'r', encoding='cp1252') as f:
            lines = [line.rstrip('\n\r') for line in f]
            return lines

def read_ttx_tab_delimited(file_path, scan_rows=10):
    """Read tab-delimited TTX file and return rows."""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            lines = [line.rstrip('\n\r') for line in f]
            return lines
    except UnicodeDecodeError:
        with open(file_path, 'r', encoding='cp1252') as f:
            lines = [line.rstrip('\n\r') for line in f]
            return lines

def detect_txt_header_row(lines, scan_rows=20):
    """
    Detect the header row in a pipe-delimited text file.
    The header is the topmost row with the same number of pipes as the majority of rows.
    """
    if not lines:
        return 0
    
    # Limit to scan_rows
    scan_lines = lines[:min(scan_rows, len(lines))]
    
    # Count pipes in each row
    pipe_counts = []
    for line in scan_lines:
        pipe_count = line.count('|')
        pipe_counts.append(pipe_count)
    
    # Find the most common pipe count
    if not pipe_counts:
        return 0
    
    # Count frequency of each pipe count
    from collections import Counter
    count_freq = Counter(pipe_counts)
    
    # Get the most common pipe count
    most_common_pipe_count = count_freq.most_common(1)[0][0]
    
    # Find the first row with this pipe count
    for i, pipe_count in enumerate(pipe_counts):
        if pipe_count == most_common_pipe_count:
            return i
    
    return 0

def detect_ttx_header_row(lines, scan_rows=20):
    """
    Detect the header row in a tab-delimited TTX file.
    The header is the topmost row with the same number of tabs as the majority of rows.
    """
    if not lines:
        return 0
    
    # Limit to scan_rows
    scan_lines = lines[:min(scan_rows, len(lines))]
    
    # Count tabs in each row
    tab_counts = []
    for line in scan_lines:
        tab_count = line.count('\t')
        tab_counts.append(tab_count)
    
    # Find the most common tab count
    if not tab_counts:
        return 0
    
    # Count frequency of each tab count
    from collections import Counter
    count_freq = Counter(tab_counts)
    
    # Get the most common tab count
    most_common_tab_count = count_freq.most_common(1)[0][0]
    
    # Find the first row with this tab count
    for i, tab_count in enumerate(tab_counts):
        if tab_count == most_common_tab_count:
            return i
    
    return 0

def read_csv_with_fallback(file_path):
    try:
        with open(file_path, newline='', encoding='utf-8') as csvfile:
            return list(csv.reader(csvfile))
    except UnicodeDecodeError:
        with open(file_path, newline='', encoding='cp1252') as csvfile:
            return list(csv.reader(csvfile))

def detect_csv_header_row(rows, scan_rows=20):
    max_score = -1
    header_row_idx = 0
    
    for i, row in enumerate(rows[:scan_rows]):
        # Score: count of non-empty, non-numeric, unique values
        non_empty = [cell for cell in row if cell not in (None, "")]
        
        # Skip rows with too few columns (likely title rows)
        if len(non_empty) < 3:
            continue
        
        string_like = [cell for cell in non_empty if isinstance(cell, str) and not is_number(cell)]
        unique = len(set(non_empty))
        
        # Base score
        score = len(non_empty) + len(string_like) + unique
        
        # Bonus: Check if followed by data rows with similar column structure
        if i < len(rows) - 1:
            next_rows_col_counts = []
            for j in range(i + 1, min(i + 6, len(rows))):
                if j < len(rows):
                    next_row_non_empty = [cell for cell in rows[j] if cell not in (None, "")]
                    if len(next_row_non_empty) > 0:
                        next_rows_col_counts.append(len(next_row_non_empty))
            
            # If the next rows have similar column counts, boost score
            if next_rows_col_counts:
                avg_next_cols = sum(next_rows_col_counts) / len(next_rows_col_counts)
                if abs(avg_next_cols - len(non_empty)) <= 3:  # Within 3 columns
                    score += 10  # Bonus for data consistency
        
        if is_likely_header(row) and score > max_score:
            max_score = score
            header_row_idx = i
    return header_row_idx

def extract_csv_metadata(file_path, relative_path=None, scan_rows=20):
    try:
        file_name = os.path.basename(file_path)
        rows = read_csv_with_fallback(file_path)
        if not rows:
            return {
                'file_name': file_name,
                'file_path': relative_path if relative_path else str(file_path),
                'total_sheets': 1,
                'sheets': [{
                    'sheet_name': 'Sheet1',
                    'warning': 'CSV file is empty.'
                }],
                'file_formula_usage': '0/0 columns'
            }
        header_row_idx = detect_csv_header_row(rows, scan_rows=scan_rows)
        header_row = rows[header_row_idx]
        if not is_likely_header(header_row):
            print(f"Warning: CSV file '{file_name}' header missing, inherit from previous headers.")
            return {
                'file_name': file_name,
                'file_path': relative_path if relative_path else str(file_path),
                'total_sheets': 1,
                'sheets': [{
                    'sheet_name': 'Sheet1',
                    'warning': 'header missing, inherit from previous headers'
                }],
                'file_formula_usage': '0/0 columns'
            }
        columns = [str(col).strip() for col in header_row if col is not None and str(col).strip() != '']
        data_rows = rows[header_row_idx+1:]
        
        # Check if sheet is empty (no data rows or all rows are empty)
        if not data_rows or all(all(cell in (None, "") for cell in row) for row in data_rows):
            print(f"Warning: CSV file '{file_name}' is empty.")
            return {
                'file_name': file_name,
                'file_path': relative_path if relative_path else str(file_path),
                'total_sheets': 1,
                'sheets': [{
                    'sheet_name': 'Sheet1',
                    'warning': 'sheet is empty'
                }],
                'file_formula_usage': '0/0 columns'
            }
        
        max_cols = max(len(row) for row in data_rows) if data_rows else 0
        if max_cols > len(columns):
            extra_columns = [f"Column_{i+1}" for i in range(len(columns), max_cols)]
            columns = columns + extra_columns
        unique_columns = make_column_names_unique(columns)
        row_count = len(data_rows)
        skipped_rows = 0
        found_data = False
        for row in data_rows:
            if all(cell in (None, "") for cell in row):
                if found_data:
                    skipped_rows += 1
            else:
                found_data = True
        import pandas as pd
        df = pd.DataFrame(data_rows, columns=unique_columns)
        top_values_types = {}
        for orig_col, unique_col in zip(columns, unique_columns):
            col_data = df[unique_col]
            empty_mask = col_data.isna() | col_data.astype(str).str.strip().isin(["", "N/A", "NULL", "NA", "null", "n/a"])
            empty_count = int(empty_mask.sum())
            total_count = int(col_data.shape[0])
            col_data_nonan = col_data.dropna()
            if col_data_nonan.empty:
                top_values_types[orig_col] = {'top_values': [], 'empty_cells': (empty_count, total_count)}
                continue
            value_counts = col_data_nonan.value_counts().head(3)
            top_values_types[orig_col] = {
                'top_values': [str(val) for val in value_counts.index],
                'empty_cells': (empty_count, total_count)
            }
        column_details = {}
        for col in columns:
            col_info = {
                'uses_formula': False
            }
            if 'top_values' in top_values_types[col] and top_values_types[col]['top_values']:
                col_info['top_values'] = top_values_types[col]['top_values']
            if 'empty_cells' in top_values_types[col]:
                empty_count, total_count = top_values_types[col]['empty_cells']
                col_info['empty_cells (blank, N/A, NULL)'] = f"{empty_count}/{total_count}"
            column_details[col] = col_info
        sheet_info = {
            'sheet_name': 'Sheet1',
            'header_row': header_row_idx + 1,  # 1-based index for consistency
            'column_count': len(columns),
            'columns': columns,
            'unique_columns': unique_columns,
            'row_count': row_count,
            'skipped_rows': skipped_rows,
            'has_pivot_table': False,
            'column_details': column_details,
            'original_header_count': len([col for col in header_row if col not in (None, "")]),
            'actual_data_columns': max_cols,
            'duplicate_columns': {},
            'sheet_formula_usage': f"0/{len(columns)} columns"
        }
        metadata = {
            'file_name': file_name,
            'file_path': relative_path if relative_path else str(file_path),
            'total_sheets': 1,
            'sheets': [sheet_info],
            'file_formula_usage': f"0/{len(columns)} columns"
        }
        return metadata
    except Exception as e:
        return {
            'file_name': os.path.basename(file_path),
            'error': str(e)
        }

def extract_csv_simple(file_path, relative_path=None):
    try:
        file_name = os.path.basename(file_path)
        rows = read_csv_with_fallback(file_path)
        if not rows:
            return {
                'file_name': file_name,
                'file_path': relative_path if relative_path else str(file_path),
                'sheets': [{
                    'sheet_name': 'Sheet1',
                    'warning': 'CSV file is empty.'
                }]
            }
        header_row_idx = detect_csv_header_row(rows, scan_rows=10)
        header_row = rows[header_row_idx]
        if not is_likely_header(header_row):
            print(f"Warning: CSV file '{file_name}' header missing or not found in first 20 rows.")
            return {
                'file_name': file_name,
                'file_path': relative_path if relative_path else str(file_path),
                'sheets': [{
                    'sheet_name': 'Sheet1',
                    'warning': 'header missing or not found in first 20 rows'
                }]
            }
        columns = [str(col).strip() for col in header_row if col is not None and str(col).strip() != '']
        
        # Check if sheet is empty (no data rows or all rows are empty)
        data_rows = rows[header_row_idx+1:]
        if not data_rows or all(all(cell in (None, "") for cell in row) for row in data_rows):
            print(f"Warning: CSV file '{file_name}' is empty.")
            return {
                'file_name': file_name,
                'file_path': relative_path if relative_path else str(file_path),
                'sheets': [{
                    'sheet_name': 'Sheet1',
                    'warning': 'sheet is empty'
                }]
            }
        
        return {
            'file_name': file_name,
            'file_path': relative_path if relative_path else str(file_path),
            'sheets': [{
                'sheet_name': 'Sheet1',
                'columns': columns
            }]
        }
    except Exception as e:
        return {
            'file_name': os.path.basename(file_path),
            'error': str(e)
        }

def extract_txt_metadata(file_path, relative_path=None, scan_rows=20):
    """Extract metadata from pipe-delimited text files."""
    try:
        file_name = os.path.basename(file_path)
        lines = read_txt_pipe_delimited(file_path, scan_rows)
        
        if not lines:
            return {
                'file_name': file_name,
                'file_path': relative_path if relative_path else str(file_path),
                'total_sheets': 1,
                'sheets': [{
                    'sheet_name': 'Sheet1',
                    'warning': 'TXT file is empty.'
                }],
                'file_formula_usage': '0/0 columns'
            }
        
        # Detect header row
        header_row_idx = detect_txt_header_row(lines, scan_rows)
        header_line = lines[header_row_idx]
        header_row = [col.strip() for col in header_line.split('|')]
        
        if not is_likely_header(header_row):
            print(f"Warning: TXT file '{file_name}' header missing, inherit from previous headers.")
            return {
                'file_name': file_name,
                'file_path': relative_path if relative_path else str(file_path),
                'total_sheets': 1,
                'sheets': [{
                    'sheet_name': 'Sheet1',
                    'warning': 'header missing, inherit from previous headers'
                }],
                'file_formula_usage': '0/0 columns'
            }
        
        columns = [str(col).strip() for col in header_row if col is not None and str(col).strip() != '']
        
        # Parse data rows
        data_rows = []
        for i in range(header_row_idx + 1, len(lines)):
            row = [col.strip() for col in lines[i].split('|')]
            data_rows.append(row)
        
        # Check if file is empty (no data rows or all rows are empty)
        if not data_rows or all(all(cell in (None, "") for cell in row) for row in data_rows):
            print(f"Warning: TXT file '{file_name}' is empty.")
            return {
                'file_name': file_name,
                'file_path': relative_path if relative_path else str(file_path),
                'total_sheets': 1,
                'sheets': [{
                    'sheet_name': 'Sheet1',
                    'warning': 'sheet is empty'
                }],
                'file_formula_usage': '0/0 columns'
            }
        
        max_cols = max(len(row) for row in data_rows) if data_rows else 0
        if max_cols > len(columns):
            extra_columns = [f"Column_{i+1}" for i in range(len(columns), max_cols)]
            columns = columns + extra_columns
        
        unique_columns = make_column_names_unique(columns)
        row_count = len(data_rows)
        
        # Count skipped rows
        skipped_rows = 0
        found_data = False
        for row in data_rows:
            if all(cell in (None, "") for cell in row):
                if found_data:
                    skipped_rows += 1
            else:
                found_data = True
        
        # Get top values and types using pandas
        df = pd.DataFrame(data_rows, columns=unique_columns)
        top_values_types = {}
        for orig_col, unique_col in zip(columns, unique_columns):
            col_data = df[unique_col]
            empty_mask = col_data.isna() | col_data.astype(str).str.strip().isin(["", "N/A", "NULL", "NA", "null", "n/a"])
            empty_count = int(empty_mask.sum())
            total_count = int(col_data.shape[0])
            col_data_nonan = col_data.dropna()
            if col_data_nonan.empty:
                top_values_types[orig_col] = {'top_values': [], 'empty_cells': (empty_count, total_count)}
                continue
            value_counts = col_data_nonan.value_counts().head(3)
            top_values_types[orig_col] = {
                'top_values': [str(val) for val in value_counts.index],
                'empty_cells': (empty_count, total_count)
            }
        
        # Build column details
        column_details = {}
        for col in columns:
            col_info = {
                'uses_formula': False
            }
            if 'top_values' in top_values_types[col] and top_values_types[col]['top_values']:
                col_info['top_values'] = top_values_types[col]['top_values']
            if 'empty_cells' in top_values_types[col]:
                empty_count, total_count = top_values_types[col]['empty_cells']
                col_info['empty_cells (blank, N/A, NULL)'] = f"{empty_count}/{total_count}"
            column_details[col] = col_info
        
        sheet_info = {
            'sheet_name': 'Sheet1',
            'header_row': header_row_idx + 1,  # 1-based index for consistency
            'column_count': len(columns),
            'columns': columns,
            'unique_columns': unique_columns,
            'row_count': row_count,
            'skipped_rows': skipped_rows,
            'has_pivot_table': False,
            'column_details': column_details,
            'original_header_count': len([col for col in header_row if col not in (None, "")]),
            'actual_data_columns': max_cols,
            'duplicate_columns': {},
            'sheet_formula_usage': f"0/{len(columns)} columns"
        }
        
        metadata = {
            'file_name': file_name,
            'file_path': relative_path if relative_path else str(file_path),
            'total_sheets': 1,
            'sheets': [sheet_info],
            'file_formula_usage': f"0/{len(columns)} columns"
        }
        return metadata
        
    except Exception as e:
        return {
            'file_name': os.path.basename(file_path),
            'error': str(e)
        }

def extract_txt_simple(file_path, relative_path=None):
    """Extract simple metadata from pipe-delimited text files."""
    try:
        file_name = os.path.basename(file_path)
        lines = read_txt_pipe_delimited(file_path, scan_rows=10)
        
        if not lines:
            return {
                'file_name': file_name,
                'file_path': relative_path if relative_path else str(file_path),
                'sheets': [{
                    'sheet_name': 'Sheet1',
                    'warning': 'TXT file is empty.'
                }]
            }
        
        # Detect header row
        header_row_idx = detect_txt_header_row(lines, scan_rows=20)
        header_line = lines[header_row_idx]
        header_row = [col.strip() for col in header_line.split('|')]
        
        if not is_likely_header(header_row):
            print(f"Warning: TXT file '{file_name}' header missing or not found in first 20 rows.")
            return {
                'file_name': file_name,
                'file_path': relative_path if relative_path else str(file_path),
                'sheets': [{
                    'sheet_name': 'Sheet1',
                    'warning': 'header missing or not found in first 20 rows'
                }]
            }
        
        columns = [str(col).strip() for col in header_row if col is not None and str(col).strip() != '']
        
        # Check if file is empty (no data rows or all rows are empty)
        data_rows = []
        for i in range(header_row_idx + 1, len(lines)):
            row = [col.strip() for col in lines[i].split('|')]
            data_rows.append(row)
        
        if not data_rows or all(all(cell in (None, "") for cell in row) for row in data_rows):
            print(f"Warning: TXT file '{file_name}' is empty.")
            return {
                'file_name': file_name,
                'file_path': relative_path if relative_path else str(file_path),
                'sheets': [{
                    'sheet_name': 'Sheet1',
                    'warning': 'sheet is empty'
                }]
            }
        
        return {
            'file_name': file_name,
            'file_path': relative_path if relative_path else str(file_path),
            'sheets': [{
                'sheet_name': 'Sheet1',
                'columns': columns
            }]
        }
    except Exception as e:
        return {
            'file_name': os.path.basename(file_path),
            'error': str(e)
        }

def extract_ttx_metadata(file_path, relative_path=None, scan_rows=20):
    """Extract metadata from tab-delimited TTX files."""
    try:
        file_name = os.path.basename(file_path)
        lines = read_ttx_tab_delimited(file_path, scan_rows)
        
        if not lines:
            return {
                'file_name': file_name,
                'file_path': relative_path if relative_path else str(file_path),
                'total_sheets': 1,
                'sheets': [{
                    'sheet_name': 'Sheet1',
                    'warning': 'TTX file is empty.'
                }],
                'file_formula_usage': '0/0 columns'
            }
        
        # Detect header row
        header_row_idx = detect_ttx_header_row(lines, scan_rows)
        header_line = lines[header_row_idx]
        header_row = [col.strip() for col in header_line.split('\t')]
        
        if not is_likely_header(header_row):
            print(f"Warning: TTX file '{file_name}' header missing, inherit from previous headers.")
            return {
                'file_name': file_name,
                'file_path': relative_path if relative_path else str(file_path),
                'total_sheets': 1,
                'sheets': [{
                    'sheet_name': 'Sheet1',
                    'warning': 'header missing, inherit from previous headers'
                }],
                'file_formula_usage': '0/0 columns'
            }
        
        columns = [str(col).strip() for col in header_row if col is not None and str(col).strip() != '']
        
        # Parse data rows
        data_rows = []
        for i in range(header_row_idx + 1, len(lines)):
            row = [col.strip() for col in lines[i].split('\t')]
            data_rows.append(row)
        
        # Check if file is empty (no data rows or all rows are empty)
        if not data_rows or all(all(cell in (None, "") for cell in row) for row in data_rows):
            print(f"Warning: TTX file '{file_name}' is empty.")
            return {
                'file_name': file_name,
                'file_path': relative_path if relative_path else str(file_path),
                'total_sheets': 1,
                'sheets': [{
                    'sheet_name': 'Sheet1',
                    'warning': 'sheet is empty'
                }],
                'file_formula_usage': '0/0 columns'
            }
        
        max_cols = max(len(row) for row in data_rows) if data_rows else 0
        if max_cols > len(columns):
            extra_columns = [f"Column_{i+1}" for i in range(len(columns), max_cols)]
            columns = columns + extra_columns
        
        unique_columns = make_column_names_unique(columns)
        row_count = len(data_rows)
        
        # Count skipped rows
        skipped_rows = 0
        found_data = False
        for row in data_rows:
            if all(cell in (None, "") for cell in row):
                if found_data:
                    skipped_rows += 1
            else:
                found_data = True
        
        # Get top values and types using pandas
        df = pd.DataFrame(data_rows, columns=unique_columns)
        top_values_types = {}
        for orig_col, unique_col in zip(columns, unique_columns):
            col_data = df[unique_col]
            empty_mask = col_data.isna() | col_data.astype(str).str.strip().isin(["", "N/A", "NULL", "NA", "null", "n/a"])
            empty_count = int(empty_mask.sum())
            total_count = int(col_data.shape[0])
            col_data_nonan = col_data.dropna()
            if col_data_nonan.empty:
                top_values_types[orig_col] = {'top_values': [], 'empty_cells': (empty_count, total_count)}
                continue
            value_counts = col_data_nonan.value_counts().head(3)
            top_values_types[orig_col] = {
                'top_values': [str(val) for val in value_counts.index],
                'empty_cells': (empty_count, total_count)
            }
        
        # Build column details
        column_details = {}
        for col in columns:
            col_info = {
                'uses_formula': False
            }
            if 'top_values' in top_values_types[col] and top_values_types[col]['top_values']:
                col_info['top_values'] = top_values_types[col]['top_values']
            if 'empty_cells' in top_values_types[col]:
                empty_count, total_count = top_values_types[col]['empty_cells']
                col_info['empty_cells (blank, N/A, NULL)'] = f"{empty_count}/{total_count}"
            column_details[col] = col_info
        
        sheet_info = {
            'sheet_name': 'Sheet1',
            'header_row': header_row_idx + 1,  # 1-based index for consistency
            'column_count': len(columns),
            'columns': columns,
            'unique_columns': unique_columns,
            'row_count': row_count,
            'skipped_rows': skipped_rows,
            'has_pivot_table': False,
            'column_details': column_details,
            'original_header_count': len([col for col in header_row if col not in (None, "")]),
            'actual_data_columns': max_cols,
            'duplicate_columns': {},
            'sheet_formula_usage': f"0/{len(columns)} columns"
        }
        
        metadata = {
            'file_name': file_name,
            'file_path': relative_path if relative_path else str(file_path),
            'total_sheets': 1,
            'sheets': [sheet_info],
            'file_formula_usage': f"0/{len(columns)} columns"
        }
        return metadata
        
    except Exception as e:
        return {
            'file_name': os.path.basename(file_path),
            'error': str(e)
        }

def extract_ttx_simple(file_path, relative_path=None):
    """Extract simple metadata from tab-delimited TTX files."""
    try:
        file_name = os.path.basename(file_path)
        lines = read_ttx_tab_delimited(file_path, scan_rows=10)
        
        if not lines:
            return {
                'file_name': file_name,
                'file_path': relative_path if relative_path else str(file_path),
                'sheets': [{
                    'sheet_name': 'Sheet1',
                    'warning': 'TTX file is empty.'
                }]
            }
        
        # Detect header row
        header_row_idx = detect_ttx_header_row(lines, scan_rows=20)
        header_line = lines[header_row_idx]
        header_row = [col.strip() for col in header_line.split('\t')]
        
        if not is_likely_header(header_row):
            print(f"Warning: TTX file '{file_name}' header missing or not found in first 20 rows.")
            return {
                'file_name': file_name,
                'file_path': relative_path if relative_path else str(file_path),
                'sheets': [{
                    'sheet_name': 'Sheet1',
                    'warning': 'header missing or not found in first 20 rows'
                }]
            }
        
        columns = [str(col).strip() for col in header_row if col is not None and str(col).strip() != '']
        
        # Check if file is empty (no data rows or all rows are empty)
        data_rows = []
        for i in range(header_row_idx + 1, len(lines)):
            row = [col.strip() for col in lines[i].split('\t')]
            data_rows.append(row)
        
        if not data_rows or all(all(cell in (None, "") for cell in row) for row in data_rows):
            print(f"Warning: TTX file '{file_name}' is empty.")
            return {
                'file_name': file_name,
                'file_path': relative_path if relative_path else str(file_path),
                'sheets': [{
                    'sheet_name': 'Sheet1',
                    'warning': 'sheet is empty'
                }]
            }
        
        return {
            'file_name': file_name,
            'file_path': relative_path if relative_path else str(file_path),
            'sheets': [{
                'sheet_name': 'Sheet1',
                'columns': columns
            }]
        }
    except Exception as e:
        return {
            'file_name': os.path.basename(file_path),
            'error': str(e)
        }

def clean_xml_content(file_path):
    """
    Read and clean XML content to handle common invalid characters.
    Returns cleaned XML content as string.
    """
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
    except UnicodeDecodeError:
        # Try with different encoding
        try:
            with open(file_path, 'r', encoding='latin-1') as f:
                content = f.read()
        except:
            with open(file_path, 'r', encoding='cp1252') as f:
                content = f.read()
    
    # Very robust cleaning approach for large, problematic XML files:
    
    # Step 0: Remove invalid XML control characters (0x00-0x08, 0x0B-0x0C, 0x0E-0x1F except 0x09, 0x0A, 0x0D)
    # These are not allowed in XML and will cause parsing errors
    # Keep: tab (0x09), newline (0x0A), carriage return (0x0D)
    invalid_chars = ''.join(chr(i) for i in range(0x00, 0x09))  # 0x00-0x08
    invalid_chars += ''.join(chr(i) for i in range(0x0B, 0x0D))  # 0x0B-0x0C
    invalid_chars += ''.join(chr(i) for i in range(0x0E, 0x20))  # 0x0E-0x1F
    for char in invalid_chars:
        content = content.replace(char, '')
    
    # Step 1: First, protect already-escaped entities by marking them temporarily
    # This prevents double-escaping
    content = content.replace('&amp;', '___AMP___')
    content = content.replace('&lt;', '___LT___')
    content = content.replace('&gt;', '___GT___')
    content = content.replace('&quot;', '___QUOT___')
    content = content.replace('&apos;', '___APOS___')
    
    # Step 2: Now escape all remaining unescaped special characters in text content
    # Valid XML tag names MUST start with letter, underscore, or colon (not digits!)
    # This pattern matches actual valid XML tags more strictly
    tag_pattern = r'(<\?[^>]+\?>|<!--.*?-->|<!DOCTYPE[^>]*>|<\/[a-zA-Z_:][a-zA-Z0-9_:.-]*>|<[a-zA-Z_:][a-zA-Z0-9_:.-]*(?:\s+[^>]*)?\s*\/?>)'
    parts = re.split(tag_pattern, content)
    
    cleaned_parts = []
    for part in parts:
        # If this part matches a valid XML tag pattern, keep it as-is
        if part and re.match(tag_pattern, part):
            cleaned_parts.append(part)
        else:
            # This is text content - escape special characters
            # Escape & (all remaining unescaped ampersands)
            part = part.replace('&', '&amp;')
            # Escape < not followed by valid tag-start characters
            part = re.sub(r'<(?![a-zA-Z_:/!?])', '&lt;', part)
            # Escape standalone >
            part = part.replace('>', '&gt;')
            cleaned_parts.append(part)
    
    content = ''.join(cleaned_parts)
    
    # Step 3: Restore the originally-escaped entities
    content = content.replace('___AMP___', '&amp;')
    content = content.replace('___LT___', '&lt;')
    content = content.replace('___GT___', '&gt;')
    content = content.replace('___QUOT___', '&quot;')
    content = content.replace('___APOS___', '&apos;')
    
    return content

def extract_xml_paths(element, path="", visited_paths=None):
    """
    Recursively extract all unique XML element paths.
    Returns a set of unique paths in the format: root, root/element, root/element/child
    """
    if visited_paths is None:
        visited_paths = set()
    
    # Clean namespace from tag if present
    tag = element.tag
    if '}' in tag:
        tag = tag.split('}', 1)[1]
    
    # Build current path
    if path:
        current_path = f"{path}/{tag}"
    else:
        current_path = tag
    
    # Add current path to set (set automatically handles uniqueness)
    visited_paths.add(current_path)
    
    # Recursively process child elements
    for child in element:
        extract_xml_paths(child, current_path, visited_paths)
    
    return visited_paths

def extract_xml_metadata(file_path, relative_path=None):
    """Extract metadata from XML files."""
    try:
        file_name = os.path.basename(file_path)
        root = None
        parse_error = None
        
        # First attempt: Parse XML file directly
        try:
            tree = ET.parse(file_path)
            root = tree.getroot()
        except ET.ParseError as e:
            parse_error = str(e)
            # Second attempt: Clean the XML content and try again
            try:
                cleaned_content = clean_xml_content(file_path)
                root = ET.fromstring(cleaned_content)
                print(f"  Note: XML file '{file_name}' had invalid characters that were cleaned automatically")
            except Exception as clean_error:
                return {
                    'file_name': file_name,
                    'file_path': relative_path if relative_path else str(file_path),
                    'error': f'XML Parse Error: {parse_error}. Attempted to clean but failed: {str(clean_error)}'
                }
        except Exception as e:
            return {
                'file_name': file_name,
                'file_path': relative_path if relative_path else str(file_path),
                'error': str(e)
            }
        
        if root is None:
            return {
                'file_name': file_name,
                'file_path': relative_path if relative_path else str(file_path),
                'error': 'Failed to parse XML file'
            }
        
        # Extract all unique element paths
        paths = extract_xml_paths(root)
        
        # Convert set to sorted list to maintain hierarchy and ensure uniqueness
        sorted_paths = sorted(list(paths))
        
        if not sorted_paths:
            return {
                'file_name': file_name,
                'file_path': relative_path if relative_path else str(file_path),
                'total_sheets': 1,
                'sheets': [{
                    'sheet_name': 'XML',
                    'warning': 'XML file is empty or has no elements'
                }],
                'file_formula_usage': '0/0 columns'
            }
        
        # Count unique elements
        element_count = len(sorted_paths)
        
        # Create metadata in the same format as other file types
        sheet_info = {
            'sheet_name': 'XML',
            'header_row': 1,
            'column_count': element_count,
            'columns': sorted_paths,
            'unique_columns': sorted_paths,
            'row_count': 0,  # XML doesn't have rows in the same sense
            'skipped_rows': 0,
            'has_pivot_table': False,
            'column_details': {path: {'uses_formula': False} for path in sorted_paths},
            'original_header_count': element_count,
            'actual_data_columns': element_count,
            'duplicate_columns': {},
            'sheet_formula_usage': f"0/{element_count} columns"
        }
        
        metadata = {
            'file_name': file_name,
            'file_path': relative_path if relative_path else str(file_path),
            'total_sheets': 1,
            'sheets': [sheet_info],
            'file_formula_usage': f"0/{element_count} columns"
        }
        
        return metadata
        
    except Exception as e:
        return {
            'file_name': os.path.basename(file_path),
            'file_path': relative_path if relative_path else str(file_path),
            'error': str(e)
        }

def extract_xml_simple(file_path, relative_path=None):
    """Extract simple metadata from XML files."""
    try:
        file_name = os.path.basename(file_path)
        root = None
        parse_error = None
        
        # First attempt: Parse XML file directly
        try:
            tree = ET.parse(file_path)
            root = tree.getroot()
        except ET.ParseError as e:
            parse_error = str(e)
            # Second attempt: Clean the XML content and try again
            try:
                cleaned_content = clean_xml_content(file_path)
                root = ET.fromstring(cleaned_content)
                print(f"  Note: XML file '{file_name}' had invalid characters that were cleaned automatically")
            except Exception as clean_error:
                return {
                    'file_name': file_name,
                    'file_path': relative_path if relative_path else str(file_path),
                    'error': f'XML Parse Error: {parse_error}. Attempted to clean but failed: {str(clean_error)}'
                }
        except Exception as e:
            return {
                'file_name': file_name,
                'file_path': relative_path if relative_path else str(file_path),
                'error': str(e)
            }
        
        if root is None:
            return {
                'file_name': file_name,
                'file_path': relative_path if relative_path else str(file_path),
                'error': 'Failed to parse XML file'
            }
        
        # Extract all unique element paths (set ensures uniqueness)
        paths = extract_xml_paths(root)
        
        # Convert set to sorted list to maintain hierarchy and ensure uniqueness
        sorted_paths = sorted(list(paths))
        
        if not sorted_paths:
            return {
                'file_name': file_name,
                'file_path': relative_path if relative_path else str(file_path),
                'sheets': [{
                    'sheet_name': 'XML',
                    'warning': 'XML file is empty or has no elements'
                }]
            }
        
        return {
            'file_name': file_name,
            'file_path': relative_path if relative_path else str(file_path),
            'sheets': [{
                'sheet_name': 'XML',
                'columns': sorted_paths
            }]
        }
        
    except Exception as e:
        return {
            'file_name': os.path.basename(file_path),
            'file_path': relative_path if relative_path else str(file_path),
            'error': str(e)
        }

def is_likely_header(header_row):
    """
    Determine if a row is likely a header row based on multiple heuristics.
    Returns False if the row looks like data rather than headers.
    """
    # Heuristic 1: if all values are numbers or empty, it's not a real header
    if not header_row or all(
        (v is None or str(v).strip() == "" or is_number(str(v)))
        for v in header_row
    ):
        return False
    
    # Find the contiguous block of meaningful headers from the start
    meaningful_headers = []
    for v in header_row:
        if v is None or str(v).strip() == "":
            break  # Stop at first empty cell
        meaningful_headers.append(str(v))
    
    # Heuristic 2: Check if we have at least 2 meaningful headers
    if len(meaningful_headers) < 2:
        return False
    
    # Heuristic 3: Check if too many of the meaningful headers are numbers
    num_count = sum(1 for v in meaningful_headers if is_number(v))
    if num_count > len(meaningful_headers) / 2:
        return False
    
    # Heuristic 4: Check for data-like patterns (IDs, codes, large numbers)
    data_like_patterns = 0
    for v in meaningful_headers:
        v_str = str(v).strip()
        # Pattern: Large numbers with many digits (likely IDs)
        if is_number(v_str) and len(v_str.replace('.', '').replace('-', '').replace('+', '')) > 6:
            data_like_patterns += 1
        # Pattern: Leading zeros (common in IDs and formatted numbers)
        elif v_str.startswith('0') and len(v_str) > 4 and any(c.isdigit() for c in v_str):
            data_like_patterns += 1
        # Pattern: Starts with + or - followed by many zeros (scientific/financial notation)
        elif re.match(r'^[+-]0{4,}', v_str):
            data_like_patterns += 1
        # Pattern: Very short alphanumeric codes without spaces (like "PF", "OT", "N")
        elif len(v_str) <= 3 and not v_str[0].isupper() or (len(v_str) <= 2 and v_str.isupper()):
            data_like_patterns += 1
        # Pattern: Mixed alphanumeric that looks like a code (e.g., "74257X599", "RE097")
        elif re.match(r'^[A-Z0-9]{5,}$', v_str) or re.match(r'^[A-Z]{2}\d+$', v_str):
            data_like_patterns += 1
    
    # If more than 40% look like data patterns, probably not a header
    if data_like_patterns > len(meaningful_headers) * 0.4:
        return False
    
    # Heuristic 5: Headers typically have spaces or are readable words
    # Count how many values look like readable text (contain lowercase, spaces, or common header words)
    readable_count = 0
    common_header_indicators = ['name', 'id', 'date', 'amount', 'type', 'code', 'number', 
                                'description', 'status', 'account', 'balance', 'total']
    for v in meaningful_headers:
        v_lower = str(v).lower()
        # Contains spaces (multi-word headers)
        if ' ' in v_lower:
            readable_count += 1
        # Contains lowercase letters (mixed case headers)
        elif any(c.islower() for c in v_lower):
            readable_count += 1
        # Matches common header words
        elif any(indicator in v_lower for indicator in common_header_indicators):
            readable_count += 1
    
    # If less than 30% look readable, probably not a header
    if readable_count < len(meaningful_headers) * 0.3:
        return False
    
    return True

def is_number(s):
    try:
        float(s)
        return True
    except Exception:
        return False

def read_patterns_from_file(pattern_file_path):
    """Read patterns from a file (one pattern per line)."""
    patterns = []
    try:
        with open(pattern_file_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#'):  # Skip empty lines and comments
                    patterns.append(line)
        return patterns
    except Exception as e:
        print(f"Error reading pattern file: {e}")
        return []

def load_config(config_path='config.yaml'):
    """Load configuration from YAML file."""
    try:
        config_file = Path(config_path)
        if not config_file.exists():
            return None
        
        with open(config_file, 'r', encoding='utf-8') as f:
            config = yaml.safe_load(f)
        
        # Check if config is enabled
        if not config or not config.get('enabled', False):
            return None
        
        return config
    except Exception as e:
        print(f"Warning: Error loading config file: {e}")
        print("Falling back to interactive mode.\n")
        return None

def validate_config(config):
    """Validate and process configuration."""
    errors = []
    
    # Validate target folder
    target_path = config.get('target_folder', {}).get('path', '')
    if not target_path:
        errors.append("Target folder path is not specified in config")
    elif not Path(target_path).exists():
        errors.append(f"Target folder does not exist: {target_path}")
    elif not Path(target_path).is_dir():
        errors.append(f"Target path is not a directory: {target_path}")
    
    # Validate extraction mode
    mode = config.get('extraction', {}).get('mode', 'full')
    if mode not in ['simple', 'full']:
        errors.append(f"Invalid extraction mode: {mode}. Must be 'simple' or 'full'")
    
    # Validate file selection mode
    selection_mode = config.get('file_selection', {}).get('mode', 'all')
    if selection_mode not in ['all', 'single', 'list']:
        errors.append(f"Invalid file selection mode: {selection_mode}")
    
    if errors:
        print("Configuration errors:")
        for error in errors:
            print(f"  - {error}")
        return False
    
    return True

def get_folders_from_config(config):
    """Get folders to scan based on configuration."""
    target_folder = Path(config['target_folder']['path'])
    scan_mode = config['target_folder'].get('scan_mode', 'all')
    specific_subfolders = config['target_folder'].get('specific_subfolders', [])
    
    # If specific subfolder names are provided, use them
    if specific_subfolders:
        folders = []
        for subfolder_name in specific_subfolders:
            subfolder_path = target_folder / subfolder_name
            if subfolder_path.exists() and subfolder_path.is_dir():
                folders.append(subfolder_path)
            else:
                print(f"Warning: Subfolder not found: {subfolder_name}")
        return folders if folders else [target_folder]
    
    # Otherwise, use scan_mode
    subfolders = [f for f in target_folder.iterdir() if f.is_dir()]
    
    if scan_mode == "all":
        return subfolders if subfolders else [target_folder]
    elif scan_mode == "root_only":
        return [target_folder]
    elif isinstance(scan_mode, list):
        # scan_mode is a list of indices
        selected_folders = []
        for idx in scan_mode:
            if 1 <= idx <= len(subfolders):
                selected_folders.append(subfolders[idx - 1])
            else:
                print(f"Warning: Invalid subfolder index: {idx}")
        return selected_folders if selected_folders else [target_folder]
    else:
        return [target_folder]

def get_patterns_from_config(config):
    """Get file filtering patterns from configuration."""
    file_filter = config.get('file_filtering', {})
    
    if not file_filter.get('enabled', False):
        return None, False
    
    # Check if patterns should be loaded from file
    pattern_file = file_filter.get('pattern_file', '')
    if pattern_file and Path(pattern_file).exists():
        patterns = read_patterns_from_file(pattern_file)
    else:
        patterns = file_filter.get('patterns', [])
    
    one_per_pattern = file_filter.get('one_per_pattern', False)
    
    return patterns if patterns else None, one_per_pattern

def read_patterns_from_file_old(pattern_file_path):
    """Read filename patterns from a file."""
    try:
        patterns = []
        with open(pattern_file_path, 'r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#'):  # Skip empty lines and comments
                    patterns.append(line)
        return patterns
    except Exception as e:
        print(f"Error reading pattern file: {e}")
        return []

def filter_files_by_patterns(files, patterns, one_per_pattern=False):
    """Filter files based on filename patterns."""
    if not one_per_pattern:
        # Original behavior: return all matching files
        filtered_files = []
        for file_path in files:
            file_name = file_path.name
            for pattern in patterns:
                # Case insensitive matching for better compatibility
                if fnmatch.fnmatch(file_name.lower(), pattern.lower()):
                    filtered_files.append(file_path)
                    break  # Stop checking other patterns for this file
        return filtered_files
    else:
        # New behavior: return one file per pattern
        filtered_files = []
        pattern_matches = {}
        
        for pattern in patterns:
            pattern_matches[pattern] = []
            for file_path in files:
                file_name = file_path.name
                if fnmatch.fnmatch(file_name.lower(), pattern.lower()):
                    pattern_matches[pattern].append(file_path)
        
        # Take the first match for each pattern
        for pattern, matches in pattern_matches.items():
            if matches:
                filtered_files.append(matches[0])  # Take first match
        
        return filtered_files, pattern_matches

def ask_for_file_filtering(total_files):
    """Ask user if they want to filter files and get filtering preferences."""
    print(f"\nFound {total_files} total files.")
    
    while True:
        choice = input("Do you want to focus on a subset of files? (y/n): ").strip().lower()
        if choice in ['y', 'yes']:
            print("\nPattern examples:")
            print("  *Sales*.xlsx          - Files containing 'Sales' in the name")
            print("  AB*.xlsx              - Files starting with 'AB'")
            print("  *Report*.txt          - TXT files containing 'Report'")
            print("  *BlueStar*.ttx        - TTX files containing 'BlueStar'")
            print("  *.xml                 - All XML files")
            print("  *.pdf                 - All PDF files")
            print("  @filename.txt         - Load patterns from a file")
            
            while True:
                pattern_input = input("\nEnter filename pattern or @filename for pattern file: ").strip()
                
                if pattern_input.startswith('@'):
                    # User specified a pattern file
                    pattern_file = pattern_input[1:]  # Remove the @ symbol
                    if not os.path.exists(pattern_file):
                        print(f"Pattern file '{pattern_file}' not found.")
                        continue
                    
                    patterns = read_patterns_from_file(pattern_file)
                    if patterns:
                        print(f"Loaded {len(patterns)} patterns from '{pattern_file}'")
                        
                        # Ask user if they want all matches or one per pattern
                        while True:
                            choice = input("\nExtract files: [1] All files matching patterns, [2] One file per pattern: ").strip()
                            if choice in ['1', 'all', 'All']:
                                return patterns, False
                            elif choice in ['2', 'one', 'One']:
                                return patterns, True
                            else:
                                print("Please enter 1 for all files or 2 for one file per pattern.")
                    else:
                        print("No valid patterns found in file.")
                        continue
                else:
                    # User entered a single pattern
                    if pattern_input:
                        return [pattern_input], False
                    else:
                        print("Please enter a valid pattern.")
                        continue
        elif choice in ['n', 'no']:
            return None, False
        else:
            print("Please enter 'y' or 'n'.")

def main():
    # Try to load configuration file
    config = load_config('config.yaml')
    
    if config:
        # Validate configuration
        if not validate_config(config):
            print("\nFalling back to interactive mode due to configuration errors.\n")
            config = None
        else:
            print("=" * 60)
            print("RUNNING IN CONFIG MODE")
            print("=" * 60)
            print(f"Config file loaded successfully from: config.yaml\n")
    
    # CONFIG MODE
    if config:
        # Get settings from config
        folders = get_folders_from_config(config)
        mode = config.get('extraction', {}).get('mode', 'full')
        scan_rows = config.get('extraction', {}).get('scan_rows', 20)
        want_excel = config.get('output', {}).get('excel', {}).get('enabled', False)
        stop_on_error = config.get('advanced', {}).get('stop_on_error', True)
        filename_prefix = config.get('output', {}).get('json', {}).get('filename_prefix', 'excel_metadata')
        
        print(f"Target folder: {config['target_folder']['path']}")
        print(f"Extraction mode: {mode}")
        print(f"Scan rows: {scan_rows}")
        print(f"Excel output: {'Yes' if want_excel else 'No'}")
        print()
        
        # Determine base folder
        if len(folders) == 1:
            base_folder = folders[0]
        else:
            base_folder = folders[0].parent
        
        # Get supported extensions
        supported_extensions = set(config.get('extraction', {}).get('supported_extensions', 
                                    ['.xlsx', '.xls', '.xlsm', '.csv', '.txt', '.ttx', '.xml']))
        
        # Collect all files
        all_files = []
        for folder in folders:
            for file_path in folder.rglob('*'):
                if file_path.is_file():
                    all_files.append(file_path)
        
        if not all_files:
            print("No files found in the selected folder(s).")
            return
        
        # Separate supported and unsupported files
        supported_files = [f for f in all_files if f.suffix.lower() in supported_extensions]
        unsupported_files = [f for f in all_files if f.suffix.lower() not in supported_extensions]
        
        print(f"Found {len(all_files)} total files:")
        print(f"  - {len(supported_files)} processable files")
        print(f"  - {len(unsupported_files)} other files\n")
        
        # Get file filtering patterns
        patterns, one_per_pattern = get_patterns_from_config(config)
        
        if patterns:
            print(f"Applying {len(patterns)} file filter pattern(s)...")
            if one_per_pattern:
                filtered_files, pattern_matches = filter_files_by_patterns(all_files, patterns, one_per_pattern=True)
                print(f"Selected one file per pattern: {len(filtered_files)} files")
            else:
                filtered_files, _ = filter_files_by_patterns(all_files, patterns, one_per_pattern=False)
                print(f"Filtered to {len(filtered_files)} files matching patterns")
            
            if not filtered_files:
                print("No files matched the specified patterns.")
                return
            
            all_files = filtered_files
            print()
        
        # File selection based on config
        file_selection = config.get('file_selection', {})
        selection_mode = file_selection.get('mode', 'all')
        
        if selection_mode == 'single':
            single_file = file_selection.get('single_file', '')
            if isinstance(single_file, int):
                # Index-based selection
                if 1 <= single_file <= len(all_files):
                    files_to_process = [all_files[single_file - 1]]
                else:
                    print(f"Error: Invalid file index: {single_file}")
                    return
            else:
                # Name-based selection
                matching_files = [f for f in all_files if f.name == single_file]
                if matching_files:
                    files_to_process = [matching_files[0]]
                else:
                    print(f"Error: File not found: {single_file}")
                    return
        elif selection_mode == 'list':
            file_indices = file_selection.get('file_indices', [])
            files_to_process = []
            for idx in file_indices:
                if 1 <= idx <= len(all_files):
                    files_to_process.append(all_files[idx - 1])
                else:
                    print(f"Warning: Invalid file index: {idx}")
            if not files_to_process:
                print("Error: No valid files selected from indices")
                return
        else:  # 'all'
            files_to_process = all_files
        
        print(f"Processing {len(files_to_process)} file(s)...\n")
        print("=" * 60)
    
    # INTERACTIVE MODE
    else:
        print("=" * 60)
        print("RUNNING IN INTERACTIVE MODE")
        print("=" * 60)
        print("(To use configuration file, set 'enabled: true' in config.yaml)\n")
        
        folders = get_folder_selection()
        # Note: get_folder_selection() now always returns a list, never None
        
        # Determine the base folder for relative path calculation
        # Get the parent folder that contains all selected folders
        if len(folders) == 1:
            base_folder = folders[0]
        else:
            # If multiple folders, use their common parent
            base_folder = folders[0].parent

        # Collect ALL files from selected folders (recursively), not just supported types
        all_files = []
        supported_extensions = {'.xlsx', '.xls', '.xlsm', '.csv', '.txt', '.ttx', '.xml'}
        
        for folder in folders:
            # Get all files (excluding directories)
            for file_path in folder.rglob('*'):
                if file_path.is_file():
                    all_files.append(file_path)
        
        if not all_files:
            message = "No files found in the selected folder(s)."
            print(message)
            return
        
        # Separate supported and unsupported files for reporting
        supported_files = [f for f in all_files if f.suffix.lower() in supported_extensions]
        unsupported_files = [f for f in all_files if f.suffix.lower() not in supported_extensions]
        
        print(f"\nFound {len(all_files)} total files:")
        print(f"  - {len(supported_files)} processable files (Excel/CSV/TXT/TTX/XML)")
        print(f"  - {len(unsupported_files)} other files (will be listed as 'didn't process')")

        # Ask user if they want to filter files based on patterns
        patterns, one_per_pattern = ask_for_file_filtering(len(all_files))
        
        if patterns:
            # Filter files based on patterns
            if one_per_pattern:
                filtered_files, pattern_matches = filter_files_by_patterns(all_files, patterns, one_per_pattern=True)
                
                if not filtered_files:
                    print("No files matched the specified patterns.")
                    return
                
                print(f"\nFound one file per pattern ({len(filtered_files)} files):")
                print("-" * 50)
                for i, file_path in enumerate(filtered_files, 1):
                    # Find which pattern matched this file
                    matched_pattern = None
                    for pattern, matches in pattern_matches.items():
                        if matches and matches[0] == file_path:
                            matched_pattern = pattern
                            break
                    print(f"{i}. {file_path} (matches: {matched_pattern})")
                
                # Ask for confirmation
                while True:
                    confirm = input("\nProceed with these files? (y/n): ").strip().lower()
                    if confirm in ['y', 'yes']:
                        break
                    elif confirm in ['n', 'no']:
                        print("Operation cancelled.")
                        return
                    else:
                        print("Please enter 'y' or 'n'.")
            else:
                filtered_files, _ = filter_files_by_patterns(all_files, patterns)
                
                if not filtered_files:
                    print("No files matched the specified patterns.")
                    return
                
                print(f"Filtered to {len(filtered_files)} files matching the patterns.")
                if len(patterns) <= 10:  # Show patterns if not too many
                    print(f"Patterns used: {', '.join(patterns)}")
                else:
                    print(f"Used {len(patterns)} patterns for filtering.")
            
            all_files = filtered_files
        
        choice = get_user_selection(all_files)
        mode = get_user_mode()  # Moved to the end of the user input sequence
        want_excel = get_excel_output_choice()  # Ask if user wants Excel output too
        stop_on_error = True  # Default for interactive mode
        filename_prefix = 'excel_metadata'
        scan_rows = 20  # Default
        files_to_process = all_files if choice == 0 else [all_files[choice - 1]]
        supported_extensions = {'.xlsx', '.xls', '.xlsm', '.csv', '.txt', '.ttx', '.xml'}
    
    # COMMON PROCESSING CODE (used by both modes)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    suffix = '_simple' if mode == 'simple' else '_full'
    output_file = f"{filename_prefix}_{timestamp}{suffix}.json"
    result = {
        'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'files': []
    }
    
    for excel_file in files_to_process:
        # Calculate relative path from base folder
        try:
            relative_path = str(Path(excel_file).relative_to(base_folder))
        except ValueError:
            # If relative_to fails, just use the file name
            relative_path = str(Path(excel_file).name)
        
        if hasattr(excel_file, 'suffix'):
            file_extension = excel_file.suffix.lower()
        else:
            # Handle cases where excel_file might not be a Path object
            excel_file_path = Path(str(excel_file))
            file_extension = excel_file_path.suffix.lower()
        
        # Check if this is an unsupported file type
        supported_extensions = {'.xlsx', '.xls', '.xlsm', '.csv', '.txt', '.ttx', '.xml'}
        if file_extension not in supported_extensions:
            # Create metadata entry for unsupported file
            file_name = os.path.basename(excel_file)
            if mode == 'simple':
                metadata = {
                    'file_name': file_name,
                    'file_path': relative_path,
                    'sheets': [{
                        'sheet_name': 'N/A',
                        'columns': ["didn't process"]
                    }]
                }
            else:  # full mode
                metadata = {
                    'file_name': file_name,
                    'file_path': relative_path,
                    'total_sheets': 1,
                    'formula_usage': 'N/A',
                    'sheets': [{
                        'sheet_name': 'N/A',
                        'columns': ["didn't process"]
                    }]
                }
            result['files'].append(metadata)
            print(f"\nSkipped (unsupported type): {excel_file}")
            continue
        
        if mode == 'simple':
            if file_extension == '.xls':
                metadata = extract_xls_simple(excel_file, relative_path)
            elif file_extension == '.csv':
                metadata = extract_csv_simple(excel_file, relative_path)
            elif file_extension == '.txt':
                metadata = extract_txt_simple(excel_file, relative_path)
            elif file_extension == '.ttx':
                metadata = extract_ttx_simple(excel_file, relative_path)
            elif file_extension == '.xml':
                metadata = extract_xml_simple(excel_file, relative_path)
            else:
                metadata = extract_excel_simple(excel_file, relative_path)
            
            # Check for errors during processing
            if 'error' in metadata:
                print(f"\nERROR processing {excel_file}: {metadata['error']}")
                if stop_on_error:
                    print("Stopping processing as requested.")
                    return
                else:
                    print("Continuing with next file...")
            
            result['files'].append(metadata)
            print(f"\nProcessed (simple): {excel_file}")
        else:
            if file_extension == '.xls':
                metadata = extract_xls_metadata(excel_file, relative_path, scan_rows)
            elif file_extension == '.csv':
                metadata = extract_csv_metadata(excel_file, relative_path, scan_rows)
            elif file_extension == '.txt':
                metadata = extract_txt_metadata(excel_file, relative_path, scan_rows)
            elif file_extension == '.ttx':
                metadata = extract_ttx_metadata(excel_file, relative_path, scan_rows)
            elif file_extension == '.xml':
                metadata = extract_xml_metadata(excel_file, relative_path, scan_rows)
            else:
                metadata = extract_excel_metadata(excel_file, relative_path, scan_rows)
            
            # Check for errors during processing
            if 'error' in metadata:
                print(f"\nERROR processing {excel_file}: {metadata['error']}")
                if stop_on_error:
                    print("Stopping processing as requested.")
                    return
                else:
                    print("Continuing with next file...")
            
            formatted_output = format_metadata(metadata)
            result['files'].append(formatted_output)
            print(f"\nProcessed: {excel_file}")
    
    # Save JSON results
    serializable_result = ensure_serializable(result)
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(serializable_result, f, indent=2)
    print(f"\nJSON report has been saved to: {output_file}")
    
    # Optionally create Excel file
    if want_excel:
        excel_file = convert_json_to_excel(output_file)
        if excel_file:
            print(f"Excel report has been saved to: {excel_file}")
        else:
            print("Failed to create Excel file.")

if __name__ == "__main__":
    main() 